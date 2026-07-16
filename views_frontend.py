# Controladores de vistas web encargados de renderizar HTML
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.db import transaction
from django.db.models import Q
from django.core.exceptions import ValidationError
from datetime import date, datetime, timedelta, time
from decimal import Decimal
from hotel.models import Habitacion, Hotel, TipoHabitacion
from huespedes.models import Huesped
from reservas.models import Reserva
from estancias.models import Estancia, CargoEstancia, Folio, Pago
from utils.auditoria import log_action
from reportes.models import registrar_auditoria
import requests
from config import roles


# ── Helpers de Rol ─────────────────────────────────────────────────────────────
def _es_admin(user):
    return roles.es_admin(user)

def _es_recepcionista(user):
    return roles.es_recepcionista(user)

def _es_housekeeping(user):
    return roles.es_housekeeping(user)

def _solo_housekeeping(user):
    """True si el usuario es SOLO housekeeping (sin admin ni recepcionista)."""
    return roles.solo_housekeeping(user)

def _acceso_denegado(request, msg='No tienes permisos para acceder a esta sección.'):
    return render(request, '403.html', {'mensaje_error': msg}, status=403)


def parse_room_gallery(raw_urls, main_url=''):
    urls = []
    for raw_url in (main_url, *raw_urls.splitlines()):
        url = raw_url.strip()
        if url and url not in urls:
            urls.append(url)
    return urls


def parse_datetime_local(value):
    if not value:
        return None
    parsed = datetime.strptime(value, '%Y-%m-%dT%H:%M')
    return timezone.make_aware(parsed)


def parse_date_local(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    value = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValidationError(f'Fecha inválida: {value}')


def calcular_cargo_salida_tardia(estancia):
    from estancias.services import detectar_late_checkout
    hotel = estancia.habitacion.hotel
    now_local = timezone.localtime(timezone.now())
    es_late, late_monto, minutos_tarde = detectar_late_checkout(estancia, hotel, now_local)
    return late_monto, minutos_tarde


def login_view(request):
    if request.user.is_authenticated:
        if _solo_housekeeping(request.user):
            return redirect('housekeeping')
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username', '')
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')[:200]
        
        from django.contrib.auth.models import User
        from reportes.models import LoginIntento
        
        user = authenticate(request, username=username, password=request.POST['password'])
        
        if user:
            intentos_fallidos = LoginIntento.contar_fallidos_recientes(user, minutos=15)
            if intentos_fallidos >= 5:
                if user.is_active:
                    user.is_active = False
                    user.save(update_fields=['is_active'])
                    log_action(
                        user=None,
                        accion="Bloqueo Automático",
                        registro_id=user.id,
                        tabla_afectada="auth_user",
                        estado_anterior="Activo",
                        estado_nuevo="Bloqueado",
                        observacion=f"Usuario '{username}' bloqueado tras 5 intentos fallidos. IP: {ip}"
                    )
                messages.error(request, 'Tu cuenta está bloqueada por múltiples intentos fallidos. Contacta al administrador.')
                return redirect('login')
            
            LoginIntento.registrar(usuario=user, ip=ip, user_agent=user_agent, exitoso=True)
            
            login(request, user)
            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])
            
            log_action(
                user=user,
                accion="Login Exitoso",
                registro_id=user.id,
                tabla_afectada="auth_user",
                estado_nuevo=f"Sesión iniciada desde IP: {ip}",
                observacion=f"Device: {user_agent}"
            )
            
            if _solo_housekeeping(user):
                return redirect('housekeeping')
            return redirect('dashboard')
        else:
            LoginIntento.registrar(usuario=None, ip=ip, user_agent=user_agent, exitoso=False)
            log_action(
                user=None,
                accion="Intento Login Fallido",
                registro_id=None,
                tabla_afectada="auth_user",
                observacion=f"Usuario: '{username}', IP: {ip}, Device: {user_agent}"
            )
            messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'login.html')


def logout_view(request):
    if request.user.is_authenticated:
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        log_action(
            user=request.user,
            accion="Logout",
            tabla_afectada="auth_user",
            observacion=f"Sesión cerrada. IP: {ip}"
        )
    logout(request)
    return redirect('login')


@login_required
def dashboard(request):
    if _solo_housekeeping(request.user):
        return redirect('housekeeping')

    habitaciones = Habitacion.objects.select_related('tipo', 'hotel').all().order_by('piso', 'numero')
    hoy = timezone.now().date()
    ahora = timezone.now()
    
    total = habitaciones.count()
    ocupadas = habitaciones.filter(estado='OCUPADA').count()
    stats = {
        'disponibles': habitaciones.filter(estado='DISPONIBLE').count(),
        'ocupadas': ocupadas,
        'reservas_hoy': 0, # Calculado más abajo
        'tasa_ocupacion': round(ocupadas / total * 100) if total else 0,
    }
    
    # Lógica de estados visuales para el plano
    from datetime import datetime, time
    checkin_time = timezone.make_aware(datetime.combine(hoy, time(14, 0)))
    reservas_hoy = Reserva.objects.filter(
        fecha_entrada=hoy, estado__in=['PENDIENTE', 'CONFIRMADA']
    ).select_related('habitacion')
    reservas_dict = {r.habitacion_id: r for r in reservas_hoy if r.habitacion_id}
    
    for hab in habitaciones:
        hab.estado_visual = hab.estado
        if hab.estado == 'OCUPADA':
            for e in hab.estancias.all():
                if e.estado == 'ACTIVA' and e.reserva.fecha_salida < hoy:
                    hab.estado_visual = 'VENCIDA'
                    break
        elif hab.estado == 'DISPONIBLE':
            r = reservas_dict.get(hab.id)
            if r:
                hab.reserva_hoy = r
                if ahora > checkin_time:
                    hab.estado_visual = 'RETRASO'
                else:
                    hab.estado_visual = 'RESERVADA'

    stats['reservas_hoy'] = sum(1 for hab in habitaciones if getattr(hab, 'estado_visual', '') in ['RESERVADA', 'RETRASO'])

    llegadas_hoy = Reserva.objects.filter(
        fecha_entrada=hoy,
        estado__in=['PENDIENTE', 'CONFIRMADA']
    ).select_related('huesped', 'habitacion')[:15]

    en_casa = Estancia.objects.filter(
        estado='ACTIVA'
    ).select_related('reserva__huesped', 'habitacion')[:15]

    salidas_hoy = Estancia.objects.filter(
        estado='ACTIVA',
        reserva__fecha_salida=hoy
    ).select_related('reserva__huesped', 'habitacion')[:15]

    # Calcular KPIs operativos de recepción (HOT-REP-001)
    from estancias.models import CargoEstancia, Folio
    from atencion.models import TicketServicio
    from decimal import Decimal

    early_checkins_hoy = CargoEstancia.objects.filter(
        concepto__icontains="Early Check-In",
        estancia__fecha_checkin__date=hoy
    ).count()

    late_checkouts_hoy = CargoEstancia.objects.filter(
        concepto__icontains="Salida Tardía",
        fecha__date=hoy
    ).count()

    # Reservas para hoy pendientes de check-in y próximas a vencer
    reservas_vencer_hoy = Reserva.objects.filter(
        fecha_entrada=hoy,
        estado__in=['PENDIENTE', 'CONFIRMADA']
    ).count()

    tickets_pendientes = TicketServicio.objects.filter(
        estado__in=['ABIERTA', 'PROCESO', 'PENDIENTE']
    ).count()

    # Saldos de folios activos (Hospedajes activos)
    active_folios = Folio.objects.filter(estancia__estado='ACTIVA')
    saldos_activos = sum(f.saldo_pendiente for f in active_folios)

    # Permiso para ver importes financieros (RN-REP-003)
    mostrar_financiero = _es_admin(request.user)

    return render(request, 'dashboard.html', {
        'habitaciones': habitaciones,
        'stats': stats,
        'llegadas_hoy': llegadas_hoy,
        'en_casa': en_casa,
        'salidas_hoy': salidas_hoy,
        'early_checkins_hoy': early_checkins_hoy,
        'late_checkouts_hoy': late_checkouts_hoy,
        'reservas_vencer_hoy': reservas_vencer_hoy,
        'tickets_pendientes': tickets_pendientes,
        'saldos_activos': saldos_activos,
        'mostrar_financiero': mostrar_financiero,
    })


@login_required
def reservas_lista(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    estado = request.GET.get('estado')
    query = request.GET.get('q', '').strip()
    reservas = Reserva.objects.select_related('huesped', 'habitacion').all().order_by('-created_at')
    if estado:
        reservas = reservas.filter(estado=estado)
    if query:
        if query.isdigit():
            reservas = reservas.filter(id=query)
        else:
            reservas = reservas.filter(
                Q(huesped__nombres__icontains=query) |
                Q(huesped__apellidos__icontains=query) |
                Q(huesped__num_doc__icontains=query)
            )

    hoy = timezone.now().date()
    llegadas_hoy = Reserva.objects.filter(fecha_entrada=hoy, estado__in=['PENDIENTE', 'CONFIRMADA'])
    en_casa = Estancia.objects.filter(estado='ACTIVA')
    salidas_hoy = Estancia.objects.filter(estado='ACTIVA', reserva__fecha_salida=hoy)

    return render(request, 'reservas/lista.html', {
        'reservas': reservas, 
        'query': query, 
        'estado': estado,
        'llegadas_hoy': llegadas_hoy,
        'en_casa': en_casa,
        'salidas_hoy': salidas_hoy
    })


@login_required
def reserva_detalle(request, reserva_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    reserva = get_object_or_404(
        Reserva.objects.select_related('huesped', 'habitacion__tipo', 'hotel'),
        id=reserva_id,
    )
    estancia = getattr(reserva, 'estancia', None)
    return render(request, 'reservas/detalle.html', {
        'reserva': reserva,
        'estancia': estancia,
    })


@login_required
def reserva_nueva(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    if request.method == 'POST':
        from django.db import transaction
        try:
            huesped_id = request.POST.get('huesped_id')
            if huesped_id:
                huesped = Huesped.objects.get(id=huesped_id)
            else:
                huesped = Huesped.objects.get(num_doc=request.POST['huesped_doc'])
            habitacion = Habitacion.objects.get(id=request.POST['habitacion'])
            modalidad = request.POST.get('modalidad', Reserva.POR_DIA)
            fecha_hora_entrada = None
            duracion_horas = request.POST.get('duracion_horas') or 0

            if modalidad == Reserva.POR_HORA:
                fecha_hora_entrada = parse_datetime_local(request.POST.get('fecha_hora_entrada'))
                if not fecha_hora_entrada:
                    raise ValidationError('Debes indicar la fecha y hora de ingreso.')
                fecha_entrada = fecha_hora_entrada.date()
                fecha_salida = (fecha_hora_entrada + timedelta(hours=float(duracion_horas or 3))).date()
            else:
                fecha_entrada = parse_date_local(request.POST.get('fecha_entrada'))
                fecha_salida = parse_date_local(
                    request.POST.get('fecha_salida') or request.POST.get('fecha_entrada')
                )
                if not fecha_entrada or not fecha_salida:
                    raise ValidationError('Debes indicar las fechas de entrada y salida.')

            with transaction.atomic():
                reserva = Reserva.objects.create(
                    hotel=habitacion.hotel,
                    huesped=huesped,
                    habitacion=habitacion,
                    fecha_entrada=fecha_entrada,
                    fecha_salida=fecha_salida,
                    fecha_hora_entrada=fecha_hora_entrada,
                    modalidad=modalidad,
                    duracion_horas=duracion_horas,
                    num_adultos=int(request.POST.get('num_adultos', 1)),
                    origen=request.POST.get('origen', 'DIRECTO'),
                    observaciones=request.POST.get('observaciones', ''),
                )
                reserva.precio_total = reserva.calcular_precio()
                reserva.save()

                # Pago anticipado opcional
                registrar_pago = request.POST.get('registrar_pago') == 'on'
                if registrar_pago:
                    monto = request.POST.get('pago_monto')
                    metodo_pago = request.POST.get('pago_metodo_pago', Pago.EFECTIVO)
                    transaccion_id = request.POST.get('pago_transaccion_id') or None

                    if not monto:
                        raise ValidationError('Debes especificar un monto para el pago anticipado.')
                    
                    monto_decimal = Decimal(monto)
                    if monto_decimal <= 0:
                        raise ValidationError('El monto del pago debe ser mayor a cero.')
                    if monto_decimal > reserva.precio_total:
                        raise ValidationError(f'El monto del pago no puede ser mayor que el total de la reserva (S/. {reserva.precio_total:.2f}).')
                    if metodo_pago != Pago.EFECTIVO and not transaccion_id:
                        raise ValidationError('Debes ingresar el ID de transacción para pagos electrónicos/bancarios.')
                    
                    pago = Pago.objects.create(
                        reserva=reserva,
                        monto=monto_decimal,
                        metodo_pago=metodo_pago,
                        transaccion_id=transaccion_id
                    )

                    # Registrar auditoría de pago
                    # removed import, using utils.auditoria.log_action
                    registrar_auditoria(
                        usuario=request.user,
                        accion="Registrar Pago Anticipado",
                        registro_id=pago.id,
                        tabla_afectada="estancias_pago",
                        estado_nuevo=f"ID: {pago.id}, Reserva: {reserva.id}, Monto: S/. {monto_decimal}, Metodo: {metodo_pago}"
                    )

                # Registrar auditoria
                # removed import, using utils.auditoria.log_action
                registrar_auditoria(
                    usuario=request.user,
                    accion="Crear Reserva",
                    registro_id=reserva.id,
                    tabla_afectada="reservas_reserva",
                    estado_nuevo=f"ID: {reserva.id}, Huesped: {reserva.huesped.nombres} {reserva.huesped.apellidos}, Hab: {reserva.habitacion.numero}, Total: S/. {reserva.precio_total}"
                )

                messages.success(request, f'Reserva #{reserva.id} creada correctamente.')
                if request.POST.get('accion') == 'checkin':
                    return redirect('reserva_checkin', reserva_id=reserva.id)
                return redirect('reservas_lista')
        except Exception as e:
            messages.error(request, f'Error al crear la reserva: {str(e)}')

    selected_habitacion_id = request.GET.get('habitacion', '')
    selected_huesped = None
    huesped_id = request.GET.get('huesped')
    if huesped_id:
        selected_huesped = Huesped.objects.filter(id=huesped_id).first()

    return render(request, 'reservas/nueva.html', {
        'huespedes': Huesped.objects.all().order_by('nombres', 'apellidos'),
        'habitaciones': Habitacion.objects.exclude(estado='MANTENIMIENTO').select_related('tipo').order_by('piso', 'numero'),
        'selected_habitacion_id': selected_habitacion_id,
        'selected_huesped': selected_huesped,
        'hora_actual': timezone.localtime().strftime('%Y-%m-%dT%H:%M'),
    })


from django.http import JsonResponse

@login_required
def api_habitaciones_disponibles(request):
    fecha_entrada_str = request.GET.get('entrada')
    fecha_salida_str = request.GET.get('salida')
    hora_entrada_str = request.GET.get('hora_entrada')
    duracion = float(request.GET.get('duracion', 3))
    modalidad = request.GET.get('modalidad', 'DIA')

    habitaciones = Habitacion.objects.exclude(estado='MANTENIMIENTO').select_related('tipo')
    
    target_entrada = None
    target_salida = None

    try:
        from datetime import time
        if modalidad == 'DIA' and fecha_entrada_str and fecha_salida_str:
            fecha_entrada = parse_date_local(fecha_entrada_str)
            fecha_salida = parse_date_local(fecha_salida_str)
            if fecha_entrada and fecha_salida:
                hotel = Hotel.objects.first()
                hora_in = hotel.hora_checkin_estandar if (hotel and hasattr(hotel, 'hora_checkin_estandar')) else time(15, 0)
                hora_out = hotel.hora_checkout_estandar if (hotel and hasattr(hotel, 'hora_checkout_estandar')) else time(12, 0)
                target_entrada = timezone.make_aware(datetime.combine(fecha_entrada, hora_in))
                target_salida = timezone.make_aware(datetime.combine(fecha_salida, hora_out))
        elif modalidad == 'HORA' and hora_entrada_str:
            fecha_hora_entrada = parse_datetime_local(hora_entrada_str)
            if fecha_hora_entrada:
                target_entrada = fecha_hora_entrada
                target_salida = fecha_hora_entrada + timedelta(hours=duracion)
    except Exception:
        pass

    if target_entrada and target_salida:
        # Check overlaps on exact datetime windows:
        # A reservation overlaps if (start1 < end2) AND (end1 > start2)
        reservas_cruzadas = Reserva.objects.filter(
            estado__in=['PENDIENTE', 'CONFIRMADA', 'CHECKIN'],
            fecha_hora_entrada__lt=target_salida,
            fecha_hora_salida__gt=target_entrada
        ).values_list('habitacion_id', flat=True)
        
        habitaciones = habitaciones.exclude(id__in=reservas_cruzadas)

    data = []
    for h in habitaciones.order_by('piso', 'numero'):
        data.append({
            'id': h.id,
            'numero': h.numero,
            'piso': h.piso,
            'estado_label': h.get_estado_display(),
            'tipo_nombre': h.tipo.nombre,
            'precio_base': str(h.tipo.precio_base)
        })
    return JsonResponse({'habitaciones': data})




@login_required
def checkin_directo(request, huesped_id=None):
    """
    HOT-HOS-001 – Escenario Alternativo: Check-in directo sin reserva previa.
    Cuando existe disponibilidad pero el huésped no tiene reserva, el recepcionista
    puede registrar el hospedaje directamente desde esta vista.
    Crea una Reserva (origen=DIRECTO) + Estancia + Folio en una transacción atómica.
    """
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)

    import estancias.services as estancia_services
    from django.db import transaction
    from huespedes.models import Huesped
    from datetime import date, timedelta

    # Habitaciones DISPONIBLES para el selector
    hotel = Hotel.objects.first()
    habitaciones_disponibles = Habitacion.objects.filter(
        estado=Habitacion.DISPONIBLE
    ).select_related('tipo', 'hotel').order_by('numero')

    # Huésped preseleccionado si viene desde búsqueda
    huesped_presel = None
    if huesped_id:
        huesped_presel = get_object_or_404(Huesped, id=huesped_id)

    if request.method == 'POST':
        hab_id = request.POST.get('habitacion_id')
        huesped_doc = request.POST.get('num_doc', '').strip()
        num_adultos = int(request.POST.get('num_adultos', 1))
        modalidad = request.POST.get('modalidad', Reserva.POR_DIA)
        observaciones = request.POST.get('observaciones', '')
        duracion_horas = Decimal(request.POST.get('duracion_horas', '3'))
        fecha_salida_str = request.POST.get('fecha_salida', '')

        # Validaciones previas al procesamiento
        if not huesped_doc:
            messages.error(request, 'El documento de identidad del huésped es obligatorio (RN-HOS-004).')
            return redirect('checkin_directo')
        if num_adultos < 1:
            messages.error(request, 'Debe registrar al menos 1 adulto (RN-HOS-005).')
            return redirect('checkin_directo')
        if not hab_id:
            messages.error(request, 'Debe seleccionar una habitación.')
            return redirect('checkin_directo')

        # Buscar huésped por documento
        try:
            huesped = Huesped.objects.get(num_doc=huesped_doc)
        except Huesped.DoesNotExist:
            messages.error(
                request,
                f'No se encontró ningún huésped con el documento "{huesped_doc}". '
                'Primero debe registrar al huésped en el sistema.'
            )
            return redirect('checkin_directo')

        try:
            habitacion = Habitacion.objects.get(id=hab_id)
        except Habitacion.DoesNotExist:
            messages.error(request, 'La habitación seleccionada no existe.')
            return redirect('checkin_directo')

        # Calcular fecha salida
        try:
            if fecha_salida_str:
                from datetime import datetime as dt_import
                fecha_salida = dt_import.strptime(fecha_salida_str, '%Y-%m-%d').date()
            else:
                fecha_salida = date.today() + timedelta(days=1)
        except ValueError:
            fecha_salida = date.today() + timedelta(days=1)

        try:
            with transaction.atomic():
                # 1. Crear Reserva DIRECTO automáticamente
                reserva = Reserva.objects.create(
                    hotel=hotel or habitacion.hotel,
                    huesped=huesped,
                    habitacion=habitacion,
                    fecha_entrada=date.today(),
                    fecha_salida=fecha_salida,
                    modalidad=modalidad,
                    duracion_horas=duracion_horas if modalidad == Reserva.POR_HORA else 0,
                    num_adultos=num_adultos,
                    estado=Reserva.CONFIRMADA,
                    origen=Reserva.DIRECTO,
                    observaciones=observaciones or 'Check-in directo sin reserva previa.',
                )
                reserva.normalizar_horario()
                reserva.precio_total = reserva.calcular_precio()
                reserva.save()

                # 2. Procesar check-in vía el servicio estándar (aplica todas las RN)
                estancia = estancia_services.procesar_checkin(
                    reserva_id=reserva.id,
                    habitacion_id=habitacion.id,
                    usuario=request.user,
                    exonerar_early=False,
                )

            messages.success(
                request,
                f'Check-in directo completado. Huésped: {huesped.nombres} {huesped.apellidos} '
                f'— Hab. {habitacion.numero} — Estancia #{estancia.id}.'
            )
            return redirect('folio', estancia_id=estancia.id)

        except ValidationError as e:
            messages.error(request, e.message if hasattr(e, 'message') else str(e))
        except Exception as e:
            messages.error(request, f'Error al registrar el hospedaje directo: {str(e)}')

        return redirect('checkin_directo')

    return render(request, 'reservas/checkin_directo.html', {
        'habitaciones': habitaciones_disponibles,
        'huesped_presel': huesped_presel,
        'today': date.today().isoformat(),
        'manana': (date.today() + timedelta(days=1)).isoformat(),
    })


@login_required
def reserva_checkin(request, reserva_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if reserva.estado not in ['PENDIENTE', 'CONFIRMADA']:
        messages.error(request, 'Esta reserva no puede hacer check-in.')
        return redirect('reservas_lista')

    import estancias.services as estancia_services
    from estancias.models import Estancia as EstanciaModel

    # ── Auto-sanar habitaciones en estado OCUPADA sin estancia activa ──────────
    # Esto corrige datos inconsistentes que quedan cuando hay errores de transacción.
    habitaciones_huerfanas = Habitacion.objects.filter(estado=Habitacion.OCUPADA).exclude(
        id__in=EstanciaModel.objects.filter(estado=EstanciaModel.ACTIVA).values_list('habitacion_id', flat=True)
    )
    if habitaciones_huerfanas.exists():
        habitaciones_huerfanas.update(estado=Habitacion.DISPONIBLE)
    # ──────────────────────────────────────────────────────────────────────────

    if request.method == 'POST':
        hab_id = request.POST.get('habitacion')
        exonerar_early = request.POST.get('exonerar_early') in ['true', 'on', 'True']
        motivo_early = request.POST.get('motivo_exoneracion_early')

        try:
            estancia = estancia_services.procesar_checkin(
                reserva_id=reserva.id,
                habitacion_id=hab_id,
                usuario=request.user,
                exonerar_early=exonerar_early,
                motivo_exoneracion_early=motivo_early
            )
            messages.success(request, f'Check-in realizado. Estancia #{estancia.id} creada.')
            return redirect('folio', estancia_id=estancia.id)
        except ValidationError as e:
            messages.error(request, e.message if hasattr(e, 'message') else str(e))
            return redirect('reserva_checkin', reserva_id=reserva.id)

    # Habitaciones disponibles para selección en check-in:
    # 1) Todas las DISPONIBLES (cualquier tipo para permitir cambio/upgrade)
    # 2) Siempre incluir la habitación propia de la reserva aunque su estado físico
    #    sea OCUPADA (puede ser dato inconsistente sin estancia activa real).
    # 3) Excluir habitaciones que tengan una estancia ACTIVA de OTRA reserva.
    from estancias.models import Estancia as EstanciaModel
    hab_con_estancia_ajena = EstanciaModel.objects.filter(
        estado=EstanciaModel.ACTIVA
    ).exclude(reserva=reserva).values_list('habitacion_id', flat=True)

    habitaciones_disponibles = list(
        Habitacion.objects.filter(estado=Habitacion.DISPONIBLE)
        .exclude(id__in=hab_con_estancia_ajena)
        .select_related('tipo', 'hotel')
        .order_by('numero')
    )
    # Siempre asegurar que la habitación pre-asignada esté en la lista
    if reserva.habitacion:
        ids_en_lista = [h.id for h in habitaciones_disponibles]
        if reserva.habitacion.id not in ids_en_lista:
            # Solo incluir si no tiene estancia activa de otra reserva
            if reserva.habitacion.id not in hab_con_estancia_ajena:
                habitaciones_disponibles = [reserva.habitacion] + habitaciones_disponibles

    # Detectar Early Check-In para avisar al recepcionista
    es_early = False
    early_monto = Decimal('0.00')
    try:
        now_local = timezone.localtime(timezone.now())
        es_early, early_monto = estancia_services.detectar_early_checkin(reserva, reserva.hotel, now_local)
    except Exception:
        pass

    return render(request, 'reservas/checkin.html', {
        'reserva': reserva,
        'habitaciones': habitaciones_disponibles,
        'es_early': es_early,
        'early_monto': early_monto,
    })


@login_required
def folio_view(request, estancia_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    estancia = get_object_or_404(Estancia, id=estancia_id)
    folio, _ = Folio.objects.get_or_create(estancia=estancia)

    # Auto-seed room charge for estancias that existed before the auto-charge logic.
    # If the folio has no cargos but there IS a precio_final, create the cargo now.
    if not estancia.cargos.exists() and estancia.precio_final and estancia.precio_final > 0:
        reserva = estancia.reserva
        if reserva.modalidad == 'HORA':
            horas = int(reserva.duracion_horas or 3)
            concepto = f'Alquiler por horas \u2013 Hab. {estancia.habitacion.numero} ({horas}h)'
        else:
            noches = (reserva.fecha_salida - reserva.fecha_entrada).days or 1
            concepto = f'Alquiler por {noches} noche{"s" if noches != 1 else ""} \u2013 Hab. {estancia.habitacion.numero}'
        CargoEstancia.objects.create(
            estancia=estancia,
            concepto=concepto,
            monto=estancia.precio_final,
            tipo='HABITACION',
        )

    folio.calcular_totales()
    cargo_tardanza, minutos_tarde = calcular_cargo_salida_tardia(estancia)
    
    from inventario.models import Producto
    productos = Producto.objects.filter(estado='ACTIVO', es_vendible=True)

    return render(request, 'estancias/folio.html', {
        'estancia': estancia,
        'folio': folio,
        'cargo_tardanza': cargo_tardanza,
        'minutos_tarde': minutos_tarde,
        'productos_inventario': productos
    })


@login_required
def agregar_cargo(request, estancia_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    if request.method == 'POST':
        try:
            import estancias.services as estancia_services
            producto_id = request.POST.get('producto_id')
            if producto_id:
                cantidad = int(request.POST.get('cantidad', 1))
                estancia_services.registrar_consumo(
                    estancia_id=estancia_id,
                    concepto="",
                    monto=Decimal('0.00'),
                    tipo=request.POST.get('tipo', 'OTRO'),
                    usuario=request.user,
                    producto_id=int(producto_id),
                    cantidad=cantidad
                )
            else:
                estancia_services.registrar_consumo(
                    estancia_id=estancia_id,
                    concepto=request.POST['concepto'],
                    monto=Decimal(request.POST['monto']),
                    tipo=request.POST.get('tipo', 'OTRO'),
                    usuario=request.user
                )
            messages.success(request, 'Cargo agregado correctamente.')
        except Exception as e:
            messages.error(request, str(e))
    return redirect('folio', estancia_id=estancia_id)


@login_required
def registrar_pago(request, estancia_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    if request.method == 'POST':
        estancia = get_object_or_404(Estancia, id=estancia_id)
        folio, _ = Folio.objects.get_or_create(estancia=estancia)
        monto = request.POST.get('monto')
        metodo_pago = request.POST.get('metodo_pago', Pago.EFECTIVO)
        transaccion_id = request.POST.get('transaccion_id') or None

        if monto:
            try:
                import estancias.services as estancia_services
                monto_decimal = Decimal(monto)
                estancia_services.registrar_pago_folio(
                    folio_id=folio.id,
                    monto=monto_decimal,
                    metodo=metodo_pago,
                    transaccion_id=transaccion_id,
                    usuario=request.user
                )
                messages.success(request, 'Pago registrado correctamente.')
            except Exception as e:
                messages.error(request, str(e))
        else:
            messages.error(request, 'Monto inválido para el pago.')
    return redirect('folio', estancia_id=estancia_id)


@login_required
def checkout_view(request, estancia_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    estancia = get_object_or_404(Estancia, id=estancia_id)
    exonerar_late = request.POST.get('exonerar_late_checkout') in ['true', 'on', 'True'] or request.GET.get('exonerar_late_checkout') in ['true', 'on', 'True']
    motivo_late = request.POST.get('motivo_exoneracion_late') or request.GET.get('motivo_exoneracion_late')
    try:
        import estancias.services as estancia_services
        estancia_services.procesar_checkout(
            estancia_id=estancia_id,
            usuario=request.user,
            exonerar_late_checkout=exonerar_late,
            motivo_exoneracion_late=motivo_late
        )
        messages.success(request, 'Check-out realizado correctamente.')
        return redirect('folio_imprimir', estancia_id=estancia.id)
    except Exception as e:
        messages.error(request, str(e))
        return redirect('folio', estancia_id=estancia_id)


@login_required
def housekeeping_view(request):
    if not _es_housekeeping(request.user):
        return _acceso_denegado(request)
    piso = request.GET.get('piso', '')
    habitaciones = Habitacion.objects.filter(estado='LIMPIEZA').select_related('tipo').order_by('piso', 'numero')
    
    if piso:
        habitaciones = habitaciones.filter(piso=piso)
        
    pisos_disponibles = Habitacion.objects.values_list('piso', flat=True).distinct().order_by('piso')

    return render(request, 'housekeeping.html', {
        'habitaciones': habitaciones,
        'piso_filtro': piso,
        'pisos': pisos_disponibles
    })


@login_required
def housekeeping_estado(request, hab_id):
    if not _es_housekeeping(request.user):
        return _acceso_denegado(request)
    if request.method == 'POST':
        hab = get_object_or_404(Habitacion, id=hab_id)
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in ['DISPONIBLE', 'LIMPIEZA', 'MANTENIMIENTO']:
            anterior = hab.estado
            hab.estado = nuevo_estado
            hab.save()
            
            # Registrar auditoría única
            # removed import, using utils.auditoria.log_action
            registrar_auditoria(
                usuario=request.user,
                accion="Housekeeping Estado Modificado",
                registro_id=hab.id,
                tabla_afectada="hotel_habitacion",
                estado_anterior=anterior,
                estado_nuevo=nuevo_estado,
                observacion=f"Cambio estado de Hab. {hab.numero} de {anterior} a {nuevo_estado}"
            )
            messages.success(request, f'Habitación {hab.numero} actualizada a {nuevo_estado}.')
    return redirect('housekeeping')


@login_required
def reportes_view(request):
    if not (_es_admin(request.user) or _es_recepcionista(request.user)):
        return _acceso_denegado(request, 'No tiene permiso para ver esta sección.')
    
    is_admin = _es_admin(request.user)
    is_recep = _es_recepcionista(request.user)
    
    import json
    from datetime import timedelta, datetime
    from django.utils import timezone
    from django.db import models
    from django.db.models import Sum, Q, Count, Avg
    from django.db.models.functions import TruncMonth
    from estancias.models import Estancia, CargoEstancia, Folio, Pago, Reembolso, HistorialHabitacionEstancia
    from reservas.models import Reserva, Huesped
    from hotel.models import Habitacion, TipoHabitacion
    from reportes.models import Auditoria
    from atencion.models import TicketServicio
    from decimal import Decimal
    
    hoy = timezone.now().date()
    periodo = request.GET.get('periodo', '')
    fecha_inicio_str = request.GET.get('fecha_inicio', '').strip()
    fecha_fin_str = request.GET.get('fecha_fin', '').strip()
    
    # 1. Parsing dates
    if fecha_inicio_str and fecha_fin_str:
        try:
            start_date = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            periodo = 'custom'
        except ValueError:
            periodo = 'mes'
    
    if not periodo or periodo not in ['hoy', 'semana', 'mes', 'anio', 'custom']:
        periodo = 'mes'
        
    if periodo == 'hoy':
        start_date = hoy
        end_date = hoy
    elif periodo == 'semana':
        start_date = hoy - timedelta(days=hoy.weekday())
        end_date = start_date + timedelta(days=6)
    elif periodo == 'anio':
        start_date = hoy.replace(month=1, day=1)
        end_date = hoy.replace(month=12, day=31)
    elif periodo == 'mes':
        start_date = hoy.replace(day=1)
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month - timedelta(days=next_month.day)
        
    dias = (end_date - start_date).days + 1
    
    # Período anterior para comparación
    if periodo == 'hoy':
        prev_start = start_date - timedelta(days=1)
        prev_end = prev_start
    elif periodo == 'semana':
        prev_start = start_date - timedelta(days=7)
        prev_end = end_date - timedelta(days=7)
    elif periodo == 'anio':
        prev_start = start_date.replace(year=start_date.year - 1)
        prev_end = end_date.replace(year=end_date.year - 1)
    else: # mes or custom
        prev_end = start_date - timedelta(days=1)
        prev_start = prev_end - timedelta(days=dias - 1)
        
    dias_prev = (prev_end - prev_start).days + 1
    
    # Permisos del rol
    active_tab = request.GET.get('tab', 'general')
    if not is_admin:
        if active_tab in ['general', 'finanzas']:
            active_tab = 'ocupacion'
            
    # Filters
    f_piso = request.GET.get('piso', '')
    f_tipo = request.GET.get('tipo_habitacion', '')
    f_estado = request.GET.get('estado', '')
    f_habitacion = request.GET.get('habitacion', '')
    
    # KPIs base
    total_habitaciones = Habitacion.objects.count()
    rooms_maint = Habitacion.objects.filter(estado='MANTENIMIENTO').count()
    hab_disponibles_periodo = max(1, (total_habitaciones - rooms_maint) * dias)
    hab_disponibles_prev = max(1, (total_habitaciones - rooms_maint) * dias_prev)
    
    # Noches vendidas
    estancias_periodo = Estancia.objects.filter(fecha_checkin__date__gte=start_date, fecha_checkin__date__lte=end_date)
    noches_actual = 0
    for e in estancias_periodo:
        if e.reserva.modalidad == 'DIA':
            noches_actual += max(1, (e.reserva.fecha_salida - e.reserva.fecha_entrada).days)
        else:
            noches_actual += 1
            
    ocupacion_actual = round((noches_actual / hab_disponibles_periodo * 100), 1) if hab_disponibles_periodo > 0 else 0.0
    
    # Periodo anterior
    estancias_prev = Estancia.objects.filter(fecha_checkin__date__gte=prev_start, fecha_checkin__date__lte=prev_end)
    noches_prev = 0
    for e in estancias_prev:
        if e.reserva.modalidad == 'DIA':
            noches_prev += max(1, (e.reserva.fecha_salida - e.reserva.fecha_entrada).days)
        else:
            noches_prev += 1
    ocupacion_prev = round((noches_prev / hab_disponibles_prev * 100), 1) if hab_disponibles_prev > 0 else 0.0
    
    # Revenue
    revenue_total = CargoEstancia.objects.filter(
        fecha__date__gte=start_date, fecha__date__lte=end_date
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    revenue_prev = CargoEstancia.objects.filter(
        fecha__date__gte=prev_start, fecha__date__lte=prev_end
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    # ADR
    revenue_hab = CargoEstancia.objects.filter(
        tipo='HABITACION', fecha__date__gte=start_date, fecha__date__lte=end_date
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    revenue_hab_prev = CargoEstancia.objects.filter(
        tipo='HABITACION', fecha__date__gte=prev_start, fecha__date__lte=prev_end
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    adr_actual = float(revenue_hab) / noches_actual if noches_actual > 0 else 0.0
    adr_prev = float(revenue_hab_prev) / noches_prev if noches_prev > 0 else 0.0
    
    # RevPAR
    revpar_actual = float(revenue_hab) / hab_disponibles_periodo if hab_disponibles_periodo > 0 else 0.0
    revpar_prev = float(revenue_hab_prev) / hab_disponibles_prev if hab_disponibles_prev > 0 else 0.0
    
    def calc_var(actual, prev):
        if prev == 0: return 100.0 if actual > 0 else 0.0
        return round(((float(actual) - float(prev)) / float(prev)) * 100, 1)
        
    variaciones = {
        'ocupacion': calc_var(ocupacion_actual, ocupacion_prev),
        'revenue': calc_var(revenue_total, revenue_prev),
        'adr': calc_var(adr_actual, adr_prev),
        'revpar': calc_var(revpar_actual, revpar_prev)
    }
    
    context = {
        'periodo': periodo,
        'start_date': start_date,
        'end_date': end_date,
        'active_tab': active_tab,
        'is_admin': is_admin,
        'is_recep': is_recep,
        'mostrar_financiero': is_admin,
        'ocupacion_actual': ocupacion_actual,
        'revenue_total': revenue_total,
        'adr_actual': adr_actual,
        'revpar_actual': revpar_actual,
        'variaciones': varivariations if 'varivariations' in locals() else variaciones,
    }
    
    # ----------------------------------------------------
    # TAB: GENERAL
    # ----------------------------------------------------
    if active_tab == 'general':
        # Gráficos - Ocupación por tipo (activas)
        estancias_activas_qs = Estancia.objects.filter(estado='ACTIVA').select_related('habitacion__tipo')
        ocupacion_por_tipo = {}
        for e in estancias_activas_qs:
            tipo = e.habitacion.tipo.nombre
            ocupacion_por_tipo[tipo] = ocupacion_por_tipo.get(tipo, 0) + 1
            
        # Gráficos - Ingresos mensuales (12 meses)
        doce_meses_atras = hoy.replace(day=1) - timedelta(days=365)
        ingresos_mensuales = CargoEstancia.objects.filter(fecha__date__gte=doce_meses_atras) \
            .annotate(mes=TruncMonth('fecha')) \
            .values('mes') \
            .annotate(total=Sum('monto')) \
            .order_by('mes')
        
        meses_labels = [i['mes'].strftime('%b %Y') for i in ingresos_mensuales]
        meses_data = [float(i['total']) for i in ingresos_mensuales]
        
        # Gráficos - Ingresos por origen (periodo actual)
        ingresos_origen = Reserva.objects.filter(
            fecha_entrada__gte=start_date, fecha_entrada__lte=end_date, estado__in=['CHECKIN', 'CHECKOUT']
        ).values('origen').annotate(total=Sum('precio_total'))
        
        origen_labels = [i['origen'] for i in ingresos_origen]
        origen_data = [float(i['total']) for i in ingresos_origen]
        
        ultimas_estancias = Estancia.objects.all().select_related('reserva__huesped', 'habitacion__tipo').order_by('-fecha_checkin')[:8]
        reservas_pendientes = Reserva.objects.filter(estado__in=['PENDIENTE', 'CONFIRMADA'], fecha_entrada__gte=hoy).order_by('fecha_entrada')[:5]
        
        # Extra KPIs
        reservas_confirmadas = Reserva.objects.filter(fecha_entrada__gte=start_date, fecha_entrada__lte=end_date, estado__in=['CONFIRMADA', 'CHECKIN', 'CHECKOUT']).count()
        reservas_canceladas = Reserva.objects.filter(fecha_entrada__gte=start_date, fecha_entrada__lte=end_date, estado='CANCELADA').count()
        estancias_activas_count = Estancia.objects.filter(estado='ACTIVA').count()
        estancias_canceladas_count = Estancia.objects.filter(fecha_checkin__date__gte=start_date, fecha_checkin__date__lte=end_date, estado='CANCELADA').count()
        
        # Duración promedio de estancia
        duracion_avg = Reserva.objects.filter(
            fecha_entrada__gte=start_date, fecha_entrada__lte=end_date, estado__in=['CHECKIN', 'CHECKOUT']
        ).aggregate(avg_d=Avg('duracion_horas'))['avg_d'] or 0.0
        
        # Reembolsos aprobados
        reembolsos_aprobados = Reembolso.objects.filter(fecha_resolucion__date__gte=start_date, fecha_resolucion__date__lte=end_date, estado='APROBADO').aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        
        # Método de pago más utilizado
        pago_popular = Pago.objects.filter(fecha__date__gte=start_date, fecha__date__lte=end_date).values('metodo_pago').annotate(cnt=Count('id')).order_by('-cnt').first()
        metodo_pago_popular = pago_popular['metodo_pago'] if pago_popular else 'Ninguno'
        
        # Tiempo promedio de limpieza
        completions = Auditoria.objects.filter(
            tabla_afectada="hotel_habitacion",
            accion__in=["Housekeeping Estado Modificado", "Habitación Estado Modificado"],
            estado_nuevo="DISPONIBLE",
            estado_anterior="LIMPIEZA",
            fecha__date__gte=start_date,
            fecha__date__lte=end_date
        )
        total_limp_minutos = 0
        count_limp = 0
        for c in completions:
            start_log = Auditoria.objects.filter(
                tabla_afectada="hotel_habitacion",
                registro_id=c.registro_id,
                estado_nuevo="LIMPIEZA",
                fecha__lt=c.fecha
            ).order_by('-fecha').first()
            if start_log:
                diff = c.fecha - start_log.fecha
                if diff.total_seconds() > 0 and diff < timedelta(hours=6):
                    total_limp_minutos += diff.total_seconds() / 60
                    count_limp += 1
        tiempo_limpieza_prom = round(total_limp_minutos / count_limp, 1) if count_limp > 0 else 0.0
        
        # Tickets pendientes
        tkt_pendientes = TicketServicio.objects.filter(estado__in=['ABIERTA', 'PROCESO', 'PENDIENTE']).count()
        
        context.update({
            'tipos_labels_json': json.dumps(list(ocupacion_por_tipo.keys())),
            'tipos_data_json': json.dumps(list(ocupacion_por_tipo.values())),
            'meses_labels_json': json.dumps(meses_labels),
            'meses_data_json': json.dumps(meses_data),
            'origen_labels_json': json.dumps(origen_labels),
            'origen_data_json': json.dumps(origen_data),
            'ultimas_estancias': ultimas_estancias,
            'reservas_pendientes': reservas_pendientes,
            'reservas_confirmadas': reservas_confirmadas,
            'reservas_canceladas': reservas_canceladas,
            'estancias_activas_count': estancias_activas_count,
            'estancias_canceladas_count': estancias_canceladas_count,
            'duracion_avg': round(duracion_avg, 1),
            'reembolsos_aprobados': reembolsos_aprobados,
            'metodo_pago_popular': metodo_pago_popular,
            'tiempo_limpieza_prom': tiempo_limpieza_prom,
            'tickets_pendientes': tkt_pendientes,
        })
        
    # ----------------------------------------------------
    # TAB: OCUPACION
    # ----------------------------------------------------
    elif active_tab == 'ocupacion':
        habs_qs = Habitacion.objects.all().select_related('tipo')
        if f_piso:
            habs_qs = habs_qs.filter(piso=f_piso)
        if f_tipo:
            habs_qs = habs_qs.filter(tipo_id=f_tipo)
        if f_estado:
            habs_qs = habs_qs.filter(estado=f_estado)
        if f_habitacion:
            habs_qs = habs_qs.filter(id=f_habitacion)
            
        hab_list = []
        tipo_count = {}
        for h in habs_qs:
            est_hab = Estancia.objects.filter(habitacion=h, fecha_checkin__date__lte=end_date)
            noches_hab = 0
            for e in est_hab:
                e_end = e.fecha_checkout.date() if e.fecha_checkout else hoy
                e_start = e.fecha_checkin.date()
                o_start = max(e_start, start_date)
                o_end = min(e_end, end_date)
                overlap = (o_end - o_start).days
                if overlap > 0:
                    noches_hab += overlap
                elif e_start >= start_date and e_start <= end_date:
                    noches_hab += 1
            pct = round((noches_hab / dias * 100), 1) if dias > 0 else 0
            hab_list.append({
                'id': h.id,
                'numero': h.numero,
                'piso': h.piso,
                'tipo': h.tipo.nombre,
                'estado': h.estado,
                'noches_ocupadas': noches_hab,
                'pct_ocupacion': pct
            })
            if noches_hab > 0:
                tipo_count[h.tipo.nombre] = tipo_count.get(h.tipo.nombre, 0) + noches_hab
                
        tipo_solicitado = max(tipo_count, key=tipo_count.get) if tipo_count else 'Ninguno'
        noches_totales_disponibles = habs_qs.count() * dias
        noches_totales_ocupadas = sum(h['noches_ocupadas'] for h in hab_list)
        tasa_ocupacion_consolidada = round((noches_totales_ocupadas / noches_totales_disponibles * 100), 1) if noches_totales_disponibles > 0 else 0.0
        
        context.update({
            'habitaciones_reporte': hab_list,
            'noches_totales_disponibles': noches_totales_disponibles,
            'noches_totales_ocupadas': noches_totales_ocupadas,
            'tasa_ocupacion_consolidada': tasa_ocupacion_consolidada,
            'tipo_solicitado': tipo_solicitado,
            'pisos': Habitacion.objects.values_list('piso', flat=True).distinct().order_by('piso'),
            'tipos_habitacion': TipoHabitacion.objects.all(),
            'habitaciones_all': Habitacion.objects.all().order_by('numero'),
            'f_piso': f_piso,
            'f_tipo': f_tipo,
            'f_estado': f_estado,
            'f_habitacion': f_habitacion,
        })
        
    # ----------------------------------------------------
    # TAB: RESERVAS
    # ----------------------------------------------------
    elif active_tab == 'reservas':
        res_qs = Reserva.objects.filter(fecha_entrada__gte=start_date, fecha_entrada__lte=end_date).select_related('huesped', 'habitacion__tipo')
        if f_tipo:
            res_qs = res_qs.filter(habitacion__tipo_id=f_tipo)
        if f_estado:
            res_qs = res_qs.filter(estado=f_estado)
            
        reservas_lista_reporte = []
        for r in res_qs:
            anticipo = r.pagos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            reemb = Reembolso.objects.filter(pago__reserva=r, estado='APROBADO').aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            retenido = Decimal('0.00')
            if r.estado == 'CANCELADA':
                retenido = max(Decimal('0.00'), anticipo - reemb)
            reservas_lista_reporte.append({
                'id': r.id,
                'guest': f"{r.huesped.nombres} {r.huesped.apellidos}",
                'room': r.habitacion.numero if r.habitacion else 'Sin asignar',
                'tipo': r.habitacion.tipo.nombre if r.habitacion else 'Sin asignar',
                'entrada': r.fecha_entrada,
                'salida': r.fecha_salida,
                'estado': r.estado,
                'precio_total': r.precio_total,
                'anticipo': anticipo,
                'reembolso': reemb,
                'retenido': retenido,
                'motivo_cancelacion': r.motivo_cancelacion or '—'
            })
            
        counts = res_qs.values('estado').annotate(cnt=Count('id'))
        cnt_dict = {c['estado']: c['cnt'] for c in counts}
        
        cancelaciones_con_reembolso_total = 0
        cancelaciones_con_reembolso_parcial = 0
        cancelaciones_sin_reembolso = 0
        monto_reembolsado_total = Decimal('0.00')
        monto_retenido_total = Decimal('0.00')
        
        for item in reservas_lista_reporte:
            if item['estado'] == 'CANCELADA':
                monto_reembolsado_total += item['reembolso']
                monto_retenido_total += item['retenido']
                if item['anticipo'] > 0:
                    if item['reembolso'] == item['anticipo']:
                        cancelaciones_con_reembolso_total += 1
                    elif item['reembolso'] > 0:
                        cancelaciones_con_reembolso_parcial += 1
                    else:
                        cancelaciones_sin_reembolso += 1
                else:
                    cancelaciones_sin_reembolso += 1
                    
        context.update({
            'reservas_lista_reporte': reservas_lista_reporte,
            'r_pendientes': cnt_dict.get('PENDIENTE', 0),
            'r_confirmadas': cnt_dict.get('CONFIRMADA', 0),
            'r_checkin': cnt_dict.get('CHECKIN', 0),
            'r_checkout': cnt_dict.get('CHECKOUT', 0),
            'r_canceladas': cnt_dict.get('CANCELADA', 0),
            'r_reembolsado': cnt_dict.get('REEMBOLSADO', 0),
            'cancelaciones_reembolso_total': cancelaciones_con_reembolso_total,
            'cancelaciones_reembolso_parcial': cancelaciones_con_reembolso_parcial,
            'cancelaciones_sin_reembolso': cancelaciones_sin_reembolso,
            'monto_reembolsado_total': monto_reembolsado_total,
            'monto_retenido_total': monto_retenido_total,
            'tipos_habitacion': TipoHabitacion.objects.all(),
            'f_tipo': f_tipo,
            'f_estado': f_estado,
        })
        
    # ----------------------------------------------------
    # TAB: ESTANCIAS
    # ----------------------------------------------------
    elif active_tab == 'estancias':
        est_qs = Estancia.objects.filter(fecha_checkin__date__gte=start_date, fecha_checkin__date__lte=end_date).select_related('reserva__huesped', 'habitacion__tipo')
        if f_tipo:
            est_qs = est_qs.filter(habitacion__tipo_id=f_tipo)
        if f_estado:
            est_qs = est_qs.filter(estado=f_estado)
            
        estancias_lista_reporte = []
        early_count = 0
        late_count = 0
        salidas_anticipadas = 0
        
        for e in est_qs:
            total_alojamiento = e.cargos.filter(tipo='HABITACION').exclude(concepto__icontains="Salida Tardía").exclude(concepto__icontains="Early Check-In").aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            total_consumos = e.cargos.filter(tipo='CONSUMO').aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            cargo_early = e.cargos.filter(concepto__icontains="Early Check-In").aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            if cargo_early > 0:
                early_count += 1
            cargo_late = e.cargos.filter(concepto__icontains="Salida Tardía").aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            if cargo_late > 0:
                late_count += 1
            is_anticipada = False
            if e.fecha_checkout and e.fecha_checkout.date() < e.reserva.fecha_salida:
                is_anticipada = True
                salidas_anticipadas += 1
            duracion_prog_horas = (e.reserva.fecha_hora_salida - e.reserva.fecha_hora_entrada).total_seconds() / 3600 if e.reserva.fecha_hora_salida and e.reserva.fecha_hora_entrada else 0
            duracion_real_horas = 0
            if e.fecha_checkin:
                checkout_time = e.fecha_checkout or timezone.now()
                duracion_real_horas = (checkout_time - e.fecha_checkin).total_seconds() / 3600
                
            folio = getattr(e, 'folio', None)
            total_pagado = folio.total_pagado if folio else Decimal('0.00')
            saldo_pendiente = folio.saldo_pendiente if folio else Decimal('0.00')
            
            estancias_lista_reporte.append({
                'id': e.id,
                'guest': f"{e.reserva.huesped.nombres} {e.reserva.huesped.apellidos}",
                'room': e.habitacion.numero,
                'tipo': e.habitacion.tipo.nombre,
                'checkin': e.fecha_checkin,
                'checkout': e.fecha_checkout,
                'checkout_prog': e.reserva.fecha_hora_salida,
                'duracion_prog': round(duracion_prog_horas, 1),
                'duracion_real': round(duracion_real_horas, 1),
                'total_alojamiento': total_alojamiento,
                'total_consumos': total_consumos,
                'cargo_early': cargo_early,
                'cargo_late': cargo_late,
                'total_pagado': total_pagado,
                'saldo_pendiente': saldo_pendiente,
                'is_anticipada': is_anticipada,
                'estado': e.estado
            })
            
        context.update({
            'estancias_lista_reporte': estancias_lista_reporte,
            'est_activas': est_qs.filter(estado='ACTIVA').count(),
            'est_finalizadas': est_qs.filter(estado='FINALIZADA').count(),
            'est_canceladas': est_qs.filter(estado='CANCELADA').count(),
            'early_count': early_count,
            'late_count': late_count,
            'salidas_anticipadas': salidas_anticipadas,
            'tipos_habitacion': TipoHabitacion.objects.all(),
            'f_tipo': f_tipo,
            'f_estado': f_estado,
        })
        
    # ----------------------------------------------------
    # TAB: FINANZAS (SOLO ADMIN)
    # ----------------------------------------------------
    elif active_tab == 'finanzas' and is_admin:
        saldo_inicial = Pago.objects.filter(fecha__lt=start_date).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        reembolsos_inicial = Reembolso.objects.filter(fecha_resolucion__lt=start_date, estado='APROBADO').aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        saldo_inicial_neto = saldo_inicial - reembolsos_inicial
        
        pagos_qs = Pago.objects.filter(fecha__date__gte=start_date, fecha__date__lte=end_date).select_related('folio__estancia__habitacion', 'reserva__huesped')
        reemb_qs = Reembolso.objects.filter(fecha_resolucion__date__gte=start_date, fecha_resolucion__date__lte=end_date, estado='APROBADO').select_related('pago__reserva__huesped', 'pago__folio__estancia__habitacion')
        
        ingresos_periodo = pagos_qs.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        reembolsos_periodo = reemb_qs.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        egresos_periodo = Decimal('0.00')
        balance_neto = ingresos_periodo - reembolsos_periodo
        
        metodos_pago_stats = pagos_qs.values('metodo_pago').annotate(total=Sum('monto')).order_by('-total')
        
        cargos_caja_qs = CargoEstancia.objects.filter(fecha__date__gte=start_date, fecha__date__lte=end_date)
        origen_cargos_stats = cargos_caja_qs.values('tipo').annotate(total=Sum('monto')).order_by('-total')
        
        movimientos = []
        for p in pagos_qs:
            refunded_amt = p.reembolsos.filter(estado='APROBADO').aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            movimientos.append({
                'tipo': 'INGRESO',
                'origen': f"Folio #{p.folio.id}" if p.folio else f"Reserva #{p.reserva.id}",
                'huesped': p.folio.estancia.reserva.huesped.nombre_completo if p.folio else p.reserva.huesped.nombre_completo,
                'metodo': p.get_metodo_pago_display(),
                'referencia': p.transaccion_id or '—',
                'fecha': p.fecha,
                'monto': p.monto,
                'reembolsado': refunded_amt
            })
        for r in reemb_qs:
            movimientos.append({
                'tipo': 'REEMBOLSO',
                'origen': f"Pago #{r.pago.id}",
                'huesped': r.pago.folio.estancia.reserva.huesped.nombre_completo if r.pago.folio else r.pago.reserva.huesped.nombre_completo,
                'metodo': r.pago.get_metodo_pago_display(),
                'referencia': f"Reembolso #{r.id}",
                'fecha': r.fecha_resolucion,
                'monto': -r.monto,
                'reembolsado': Decimal('0.00')
            })
        movimientos.sort(key=lambda x: x['fecha'], reverse=True)
        
        context.update({
            'saldo_inicial_neto': saldo_inicial_neto,
            'ingresos_periodo': ingresos_periodo,
            'egresos_periodo': egresos_periodo,
            'reembolsos_periodo': reembolsos_periodo,
            'balance_neto': balance_neto,
            'metodos_pago_stats': metodos_pago_stats,
            'origen_cargos_stats': origen_cargos_stats,
            'movimientos': movimientos,
        })
        
    # ----------------------------------------------------
    # TAB: CLIENTES
    # ----------------------------------------------------
    elif active_tab == 'clientes':
        huespedes_qs = Huesped.objects.annotate(
            total_bookings=Count('reservas', distinct=True),
            total_estancias=Count('reservas__estancia', distinct=True)
        )
        huespedes_lista_reporte = []
        for h in huespedes_qs:
            pagos_h = Pago.objects.filter(Q(folio__estancia__reserva__huesped=h) | Q(reserva__huesped=h))
            gasto_total = pagos_h.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            reemb_h = Reembolso.objects.filter(Q(pago__folio__estancia__reserva__huesped=h) | Q(pago__reserva__huesped=h), estado='APROBADO')
            reembolso_total = reemb_h.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            gasto_neto = gasto_total - reembolso_total
            cancelaciones = h.reservas.filter(estado='CANCELADA').count()
            noches_total = 0
            estancias_h = Estancia.objects.filter(reserva__huesped=h)
            for e in estancias_h:
                if e.reserva.modalidad == 'DIA':
                    noches_total += max(1, (e.reserva.fecha_salida - e.reserva.fecha_entrada).days)
                else:
                    noches_total += 1
            if h.total_estancias > 0 or gasto_neto > 0:
                huespedes_lista_reporte.append({
                    'id': h.id,
                    'nombre': h.nombre_completo,
                    'doc': f"{h.tipo_doc} {h.num_doc}",
                    'telefono': h.telefono or '—',
                    'bookings_count': h.total_bookings,
                    'estancias_count': h.total_estancias,
                    'cancelaciones_count': cancelaciones,
                    'noches_hospedadas': noches_total,
                    'gasto_neto': gasto_neto,
                    'promedio_estancia': round(noches_total / h.total_estancias, 1) if h.total_estancias > 0 else 0
                })
        huespedes_lista_reporte.sort(key=lambda x: x['gasto_neto'], reverse=True)
        context.update({
            'huespedes_lista_reporte': huespedes_lista_reporte[:30],
        })
        
    # ----------------------------------------------------
    # TAB: HOUSEKEEPING
    # ----------------------------------------------------
    elif active_tab == 'housekeeping':
        limpiezas_aud = Auditoria.objects.filter(
            tabla_afectada="hotel_habitacion",
            accion__in=["Housekeeping Estado Modificado", "Habitación Estado Modificado"],
            estado_nuevo="DISPONIBLE",
            estado_anterior="LIMPIEZA",
            fecha__date__gte=start_date,
            fecha__date__lte=end_date
        ).select_related('usuario')
        
        limpiezas_lista = []
        user_counts = {}
        total_minutos = 0
        count_limp = 0
        
        for c in limpiezas_aud:
            start_log = Auditoria.objects.filter(
                tabla_afectada="hotel_habitacion",
                registro_id=c.registro_id,
                estado_nuevo="LIMPIEZA",
                fecha__lt=c.fecha
            ).order_by('-fecha').first()
            duracion = '—'
            if start_log:
                diff = c.fecha - start_log.fecha
                if diff.total_seconds() > 0 and diff < timedelta(hours=6):
                    mins = int(diff.total_seconds() // 60)
                    duracion = f"{mins} min"
                    total_minutos += mins
                    count_limp += 1
            limpiezas_lista.append({
                'habitacion_id': c.registro_id,
                'fecha': c.fecha,
                'usuario': c.usuario.username if c.usuario else 'Sistema',
                'duracion': duracion,
                'estado_anterior': c.estado_anterior,
                'estado_nuevo': c.estado_nuevo
            })
            u_name = c.usuario.username if c.usuario else 'Sistema'
            user_counts[u_name] = user_counts.get(u_name, 0) + 1
            
        tiempo_limpieza_prom = round(total_minutos / count_limp, 1) if count_limp > 0 else 0
        status_counts = Habitacion.objects.values('estado').annotate(cnt=Count('id'))
        status_dict = {s['estado']: s['cnt'] for s in status_counts}
        
        context.update({
            'limpiezas_lista': limpiezas_lista[:50],
            'tiempo_limpieza_prom': tiempo_limpieza_prom,
            'limpiezas_por_usuario': user_counts,
            'hab_disponibles': status_dict.get('DISPONIBLE', 0),
            'hab_ocupadas': status_dict.get('OCUPADA', 0),
            'hab_limpieza': status_dict.get('LIMPIEZA', 0),
            'hab_mantenimiento': status_dict.get('MANTENIMIENTO', 0),
        })
        
    # ----------------------------------------------------
    # TAB: ATENCION
    # ----------------------------------------------------
    elif active_tab == 'atencion':
        tickets_qs = TicketServicio.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).select_related('estancia__habitacion')
        by_state = tickets_qs.values('estado').annotate(cnt=Count('id'))
        state_dict = {s['estado']: s['cnt'] for s in by_state}
        by_priority = tickets_qs.values('prioridad').annotate(cnt=Count('id'))
        priority_dict = {p['prioridad']: p['cnt'] for p in by_priority}
        by_cat = tickets_qs.values('categoria').annotate(cnt=Count('id')).order_by('-cnt')
        
        tickets_resueltos = tickets_qs.filter(estado__in=['RESUELTA', 'CERRADA'], resolved_at__isnull=False)
        total_res_minutos = 0
        count_res = 0
        for t in tickets_resueltos:
            diff = t.resolved_at - t.created_at
            if diff.total_seconds() > 0:
                total_res_minutos += diff.total_seconds() // 60
                count_res += 1
        tiempo_resolucion_prom = round(total_res_minutos / count_res, 1) if count_res > 0 else 0
        
        tickets_pendientes_list = []
        pending_tkts = tickets_qs.filter(estado__in=['ABIERTA', 'PROCESO', 'PENDIENTE']).order_by('-created_at')
        for t in pending_tkts:
            diff = timezone.now() - t.created_at
            horas = int(diff.total_seconds() // 3600)
            mins = int((diff.total_seconds() % 3600) // 60)
            elapsed = f"{horas}h {mins}m" if horas > 0 else f"{mins} min"
            tickets_pendientes_list.append({
                'id': t.id,
                'numero': t.numero_atencion,
                'habitacion': t.estancia.habitacion.numero,
                'categoria': t.get_categoria_display(),
                'prioridad': t.prioridad,
                'estado': t.get_estado_display(),
                'descripcion': t.descripcion,
                'responsable': t.get_responsable_display(),
                'tiempo_transcurrido': elapsed,
                'fecha': t.created_at
            })
            
        context.update({
            'tickets_pendientes_list': tickets_pendientes_list,
            'tiempo_resolucion_prom': tiempo_resolucion_prom,
            'tkt_abiertos': state_dict.get('ABIERTA', 0),
            'tkt_proceso': state_dict.get('PROCESO', 0),
            'tkt_pendientes': state_dict.get('PENDIENTE', 0),
            'tkt_resueltos': state_dict.get('RESUELTA', 0) + state_dict.get('CERRADA', 0),
            'tkt_alta_urgente': priority_dict.get('ALTA', 0) + priority_dict.get('URGENTE', 0),
            'tkt_prioridades': priority_dict,
            'tkt_categorias': by_cat,
        })
        
    return render(request, 'reportes/dashboard.html', context)


@login_required
def reservas_calendario(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    from datetime import timedelta
    from django.utils import timezone
    
    hoy = timezone.now().date()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    if fecha_inicio_str:
        try:
            from datetime import datetime
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        except:
            fecha_inicio = hoy
    else:
        fecha_inicio = hoy

    dias = [fecha_inicio + timedelta(days=i) for i in range(30)]
    fecha_fin = dias[-1]
    
    habitaciones = Habitacion.objects.select_related('tipo').all().order_by('piso', 'numero')
    
    tipo_id = request.GET.get('tipo')
    if tipo_id:
        habitaciones = habitaciones.filter(tipo_id=tipo_id)
        
    reservas = Reserva.objects.filter(
        fecha_entrada__lte=fecha_fin,
        fecha_salida__gte=fecha_inicio,
        estado__in=['PENDIENTE', 'CONFIRMADA', 'CHECKIN']
    ).select_related('huesped', 'habitacion')
    
    eventos = []
    for r in reservas:
        if r.habitacion:
            inicio = max(r.fecha_entrada, fecha_inicio)
            fin = min(r.fecha_salida, fecha_fin)
            dias_reserva = (fin - inicio).days
            offset = (inicio - fecha_inicio).days
            
            color = '#f59e0b' if r.estado == 'PENDIENTE' else '#3b82f6' if r.estado == 'CONFIRMADA' else '#10b981'
            
            eventos.append({
                'id': r.id,
                'habitacion_id': r.habitacion.id,
                'huesped': f"{r.huesped.nombres} {r.huesped.apellidos}",
                'offset': offset,
                'width': dias_reserva or 1,
                'color': color,
                'estado': r.estado
            })
            
    tipos = TipoHabitacion.objects.all()
    
    return render(request, 'reservas/calendario.html', {
        'dias': dias,
        'habitaciones': habitaciones,
        'eventos': eventos,
        'tipos': tipos,
        'fecha_inicio': fecha_inicio,
        'tipo_id': int(tipo_id) if tipo_id else ''
    })


@login_required
def consultar_disponibilidad(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    
    fecha_entrada_str = request.GET.get('fecha_entrada')
    fecha_salida_str = request.GET.get('fecha_salida')
    tipo_hab = request.GET.get('tipo_habitacion')
    capacidad = request.GET.get('capacidad')
    
    habitaciones = Habitacion.objects.exclude(estado='MANTENIMIENTO').select_related('tipo')
    
    if tipo_hab:
        habitaciones = habitaciones.filter(tipo_id=tipo_hab)
    if capacidad:
        habitaciones = habitaciones.filter(tipo__capacidad__gte=int(capacidad))
    
    if fecha_entrada_str and fecha_salida_str:
        try:
            from datetime import datetime, time
            fecha_entrada = parse_date_local(fecha_entrada_str)
            fecha_salida = parse_date_local(fecha_salida_str)
            
            hotel = Hotel.objects.first()
            hora_in = hotel.hora_checkin_estandar if (hotel and hasattr(hotel, 'hora_checkin_estandar')) else time(15, 0)
            hora_out = hotel.hora_checkout_estandar if (hotel and hasattr(hotel, 'hora_checkout_estandar')) else time(12, 0)
            target_entrada = timezone.make_aware(datetime.combine(fecha_entrada, hora_in))
            target_salida = timezone.make_aware(datetime.combine(fecha_salida, hora_out))
            
            reservados = Reserva.objects.filter(
                estado__in=['PENDIENTE', 'CONFIRMADA', 'CHECKIN'],
                fecha_hora_entrada__lt=target_salida,
                fecha_hora_salida__gt=target_entrada
            ).values_list('habitacion_id', flat=True)
            
            habitaciones = habitaciones.exclude(id__in=reservados)
        except:
            pass
    
    tipos = TipoHabitacion.objects.all()
    
    return render(request, 'reservas/disponibilidad.html', {
        'habitaciones': habitaciones,
        'tipos': tipos,
        'fecha_entrada': fecha_entrada_str,
        'fecha_salida': fecha_salida_str,
        'tipo_hab': int(tipo_hab) if tipo_hab else '',
        'capacidad': capacidad,
    })


@login_required
def huespedes_lista(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    query = request.GET.get('q', '')
    huespedes = Huesped.objects.all().order_by('-created_at')
    if query:
        huespedes = huespedes.filter(
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(num_doc__icontains=query)
        )
    return render(request, 'huespedes/lista.html', {'huespedes': huespedes, 'query': query})


@login_required
def huesped_nuevo(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    next_url = request.GET.get('next') or request.POST.get('next') or ''
    if request.method == 'POST':
        try:
            huesped = Huesped.objects.create(
                tipo_doc=request.POST['tipo_doc'],
                num_doc=request.POST['num_doc'],
                nombres=request.POST['nombres'],
                apellidos=request.POST['apellidos'],
                email=request.POST.get('email') or None,
                telefono=request.POST.get('telefono') or None,
                nacionalidad=request.POST.get('nacionalidad', 'Peruana'),
            )
            messages.success(request, 'Huésped registrado correctamente.')
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                separator = '&' if '?' in next_url else '?'
                return redirect(f'{next_url}{separator}huesped={huesped.id}')
            return redirect('huespedes_lista')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return render(request, 'huespedes/form.html', {
        'next': next_url,
        'initial_query': request.GET.get('q', ''),
    })


@login_required
def huesped_editar(request, huesped_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    huesped = get_object_or_404(Huesped, id=huesped_id)
    if request.method == 'POST':
        try:
            huesped.tipo_doc = request.POST['tipo_doc']
            huesped.num_doc = request.POST['num_doc']
            huesped.nombres = request.POST['nombres']
            huesped.apellidos = request.POST['apellidos']
            huesped.email = request.POST.get('email') or None
            huesped.telefono = request.POST.get('telefono') or None
            huesped.nacionalidad = request.POST.get('nacionalidad', 'Peruana')
            huesped.save()
            messages.success(request, 'Huésped actualizado correctamente.')
            return redirect('huespedes_lista')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return render(request, 'huespedes/form.html', {
        'huesped': huesped,
        'initial_query': '',
        'next': '',
    })


@login_required
def exportar_huespedes_excel(request):
    if not _es_admin(request.user):
        return _acceso_denegado(request, 'Solo los administradores pueden exportar datos.')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from huespedes.models import Huesped
    from django.db.models import Count, Sum
    
    huespedes = Huesped.objects.annotate(
        total_estancias=Count('reservas__estancia', distinct=True),
        total_pagado=Sum('reservas__estancia__folio__pagos__monto')
    ).order_by('id')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Huéspedes"
    
    headers = ["#", "Nombre completo", "DNI", "Teléfono", "Email", "Fecha de registro", "Total estancias", "Total pagado"]
    ws.append(headers)
    
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="1E2433", end_color="1E2433", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for h in huespedes:
        fecha_reg = h.created_at.strftime('%Y-%m-%d') if hasattr(h, 'created_at') and h.created_at else "N/A"
        total_p = h.total_pagado or 0.0
        ws.append([
            h.id,
            f"{h.nombres} {h.apellidos}",
            h.num_doc,
            h.telefono or "-",
            h.email or "-",
            fecha_reg,
            h.total_estancias,
            float(total_p)
        ])
        
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 2)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="huespedes_hotelsystem.xlsx"'
    wb.save(response)
    return response


@login_required
def habitaciones_lista(request):
    if not _es_recepcionista(request.user) and not _es_housekeeping(request.user):
        return _acceso_denegado(request)
    from django.utils import timezone
    from datetime import datetime, time
    hoy = timezone.now().date()
    ahora = timezone.now()
    
    estado_filtro = request.GET.get('estado', '')
    vista = request.GET.get('vista', 'grid')
    
    habitaciones = Habitacion.objects.select_related('tipo', 'hotel').prefetch_related('estancias__reserva').all().order_by('piso', 'numero')
    if estado_filtro:
        habitaciones = habitaciones.filter(estado=estado_filtro)
        
    pisos = habitaciones.values_list('piso', flat=True).distinct().order_by('piso')
    
    # Reservas de hoy para calculos
    reservas_hoy = Reserva.objects.filter(
        fecha_entrada=hoy, estado__in=['PENDIENTE', 'CONFIRMADA']
    ).select_related('habitacion')
    reservas_dict = {r.habitacion_id: r for r in reservas_hoy if r.habitacion_id}
    
    checkin_time = timezone.make_aware(datetime.combine(hoy, time(14, 0)))

    for hab in habitaciones:
        hab.estado_visual = hab.estado
        if hab.estado == 'OCUPADA':
            for e in hab.estancias.all():
                if e.estado == 'ACTIVA' and e.reserva.fecha_salida < hoy:
                    hab.estado_visual = 'VENCIDA'
                    break
        elif hab.estado == 'DISPONIBLE':
            r = reservas_dict.get(hab.id)
            if r:
                hab.reserva_hoy = r
                if ahora > checkin_time:
                    hab.estado_visual = 'RETRASO'
                else:
                    hab.estado_visual = 'RESERVADA'
                    
    stats = {
        'disponible': sum(1 for h in habitaciones if h.estado_visual == 'DISPONIBLE'),
        'ocupada': sum(1 for h in habitaciones if h.estado_visual == 'OCUPADA'),
        'limpieza': sum(1 for h in habitaciones if h.estado_visual == 'LIMPIEZA'),
        'mantenimiento': sum(1 for h in habitaciones if h.estado_visual == 'MANTENIMIENTO'),
        'vencida': sum(1 for h in habitaciones if h.estado_visual == 'VENCIDA'),
        'retraso': sum(1 for h in habitaciones if h.estado_visual == 'RETRASO'),
    }
    
    return render(request, 'habitaciones/lista.html', {
        'habitaciones': habitaciones,
        'estado_filtro': estado_filtro,
        'vista': vista,
        'stats': stats,
        'pisos': pisos,
    })


@login_required
def habitacion_nueva(request):
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos para crear habitaciones.')
        return redirect('habitaciones_lista')

    if request.method == 'POST':
        try:
            hotel = Hotel.objects.get(id=request.POST['hotel'])
            tipo = TipoHabitacion.objects.get(id=request.POST['tipo'])
            Habitacion.objects.create(
                hotel=hotel,
                tipo=tipo,
                numero=request.POST['numero'],
                piso=request.POST['piso'],
                estado=request.POST.get('estado', 'DISPONIBLE'),
                imagen_url=request.POST.get('imagen_url') or None,
                imagenes_urls=parse_room_gallery(
                    request.POST.get('imagenes_urls', ''),
                    request.POST.get('imagen_url') or '',
                ),
            )
            messages.success(request, 'Habitación creada correctamente.')
            return redirect('habitaciones_lista')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return render(request, 'habitaciones/form.html', {
        'hoteles': Hotel.objects.all(),
        'tipos': TipoHabitacion.objects.all(),
    })


@login_required
def habitacion_editar(request, hab_id):
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos para editar habitaciones.')
        return redirect('habitaciones_lista')

    hab = get_object_or_404(Habitacion, id=hab_id)
    if request.method == 'POST':
        try:
            hab.hotel = Hotel.objects.get(id=request.POST['hotel'])
            hab.tipo = TipoHabitacion.objects.get(id=request.POST['tipo'])
            hab.numero = request.POST['numero']
            hab.piso = request.POST['piso']
            hab.estado = request.POST.get('estado', hab.estado)
            hab.imagen_url = request.POST.get('imagen_url') or None
            hab.imagenes_urls = parse_room_gallery(
                request.POST.get('imagenes_urls', ''),
                request.POST.get('imagen_url') or '',
            )
            hab.save()
            messages.success(request, 'Habitación actualizada correctamente.')
            return redirect('habitaciones_lista')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return render(request, 'habitaciones/form.html', {
        'habitacion': hab,
        'hoteles': Hotel.objects.all(),
        'tipos': TipoHabitacion.objects.all(),
    })


@login_required
def estancias_lista(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    from django.utils import timezone
    from django.db.models import Q
    
    estado = request.GET.get('estado', 'ACTIVA')
    q = request.GET.get('q', '').strip()
    
    estancias = Estancia.objects.select_related(
        'reserva__huesped', 'habitacion__tipo'
    ).order_by('-fecha_checkin')
    
    if estado != 'TODAS':
        estancias = estancias.filter(estado=estado)
        
    if q:
        estancias = estancias.filter(
            Q(reserva__huesped__nombres__icontains=q) | 
            Q(reserva__huesped__apellidos__icontains=q) |
            Q(reserva__huesped__num_doc__icontains=q)
        )

    for e in estancias:
        if e.fecha_checkout:
            e.dias_estancia = (e.fecha_checkout.date() - e.fecha_checkin.date()).days
        else:
            e.dias_estancia = (timezone.now().date() - e.fecha_checkin.date()).days or 1

    return render(request, 'estancias/lista.html', {'estancias': estancias})


@login_required
def usuarios_lista(request):
    from django.contrib.auth.models import User, Group
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos para acceder a la gestión de usuarios.')
        return redirect('dashboard')
    
    query = request.GET.get('q', '')
    usuarios = User.objects.prefetch_related('groups').all().order_by('-date_joined')
    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
    return render(request, 'usuarios/lista.html', {'usuarios': usuarios, 'query': query})

def error_403(request, exception=None):
    return render(request, '403.html', {'mensaje_error': 'La página denegó el acceso por falta de permisos.'}, status=403)

def error_404(request, exception=None):
    return render(request, '404.html', status=404)

@login_required
def reserva_imprimir_ficha(request, reserva_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    reserva = get_object_or_404(Reserva.objects.select_related('huesped', 'habitacion__tipo', 'hotel'), id=reserva_id)
    return render(request, 'reservas/imprimir_ficha.html', {'reserva': reserva})

@login_required
def folio_imprimir(request, estancia_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    estancia = get_object_or_404(Estancia, id=estancia_id)
    folio, _ = Folio.objects.get_or_create(estancia=estancia)
    folio.calcular_totales()
    return render(request, 'estancias/imprimir_folio.html', {'estancia': estancia, 'folio': folio})


@login_required
def usuario_editar(request, user_id):
    from django.contrib.auth.models import User, Group
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos para editar usuarios.')
        return redirect('dashboard')
        
    usuario_edit = get_object_or_404(User, id=user_id)
    grupos_disponibles = Group.objects.all()
    
    if request.method == 'POST':
        try:
            cambios = []
            if usuario_edit.first_name != request.POST.get('nombres', usuario_edit.first_name):
                cambios.append(f"Nombre: {usuario_edit.first_name} → {request.POST.get('nombres')}")
            if usuario_edit.last_name != request.POST.get('apellidos', usuario_edit.last_name):
                cambios.append(f"Apellido: {usuario_edit.last_name} → {request.POST.get('apellidos')}")
            if usuario_edit.email != request.POST.get('email', usuario_edit.email):
                cambios.append(f"Email: {usuario_edit.email} → {request.POST.get('email')}")
            
            usuario_edit.first_name = request.POST.get('nombres', usuario_edit.first_name)
            usuario_edit.last_name = request.POST.get('apellidos', usuario_edit.last_name)
            usuario_edit.email = request.POST.get('email', usuario_edit.email)
            usuario_edit.is_active = request.POST.get('is_active') == 'on'
            
            # Gestión de roles
            rol_id = request.POST.get('rol')
            usuario_edit.groups.clear()
            if rol_id:
                grupo = Group.objects.get(id=rol_id)
                usuario_edit.groups.add(grupo)
                cambios.append(f"Rol asignado: {grupo.name}")
                
            usuario_edit.save()
            
            registrar_auditoria(
                usuario=request.user,
                accion="Usuario Editado",
                registro_id=usuario_edit.id,
                tabla_afectada="auth_user",
                estado_nuevo=f"Usuario '{usuario_edit.username}' actualizado",
                observacion=f"Campos modificados: {', '.join(cambios)}" if cambios else "Sin cambios significativos"
            )
            
            messages.success(request, f'Usuario {usuario_edit.username} actualizado correctamente.')
            return redirect('usuarios_lista')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
            
    # Obtener el rol actual
    rol_actual = usuario_edit.groups.first()
    
    return render(request, 'usuarios/form.html', {
        'usuario_edit': usuario_edit,
        'grupos_disponibles': grupos_disponibles,
        'rol_actual': rol_actual,
    })


@login_required
def usuario_nuevo(request):
    from django.contrib.auth.models import User, Group
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos para crear usuarios.')
        return redirect('dashboard')
        
    grupos_disponibles = Group.objects.all()
    
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            password = request.POST.get('password')
            
            if User.objects.filter(username=username).exists():
                raise Exception('Ese nombre de usuario ya está en uso.')
                
            nuevo_user = User.objects.create_user(
                username=username,
                password=password,
                first_name=request.POST.get('nombres', ''),
                last_name=request.POST.get('apellidos', ''),
                email=request.POST.get('email', ''),
            )
            nuevo_user.is_active = request.POST.get('is_active') == 'on'
            
            rol_id = request.POST.get('rol')
            if rol_id:
                grupo = Group.objects.get(id=rol_id)
                nuevo_user.groups.add(grupo)
                
            nuevo_user.save()
            
            registrar_auditoria(
                usuario=request.user,
                accion="Usuario Creado",
                registro_id=nuevo_user.id,
                tabla_afectada="auth_user",
                estado_nuevo=f"Usuario '{username}' creado. Rol: {rol_id}",
                observacion=f"Email: {nuevo_user.email}, Activo: {nuevo_user.is_active}"
            )
            
            messages.success(request, f'Usuario {nuevo_user.username} creado correctamente.')
            return redirect('usuarios_lista')
        except Exception as e:
            messages.error(request, f'Error al crear: {str(e)}')
            
    return render(request, 'usuarios/nuevo.html', {
        'grupos_disponibles': grupos_disponibles,
    })


@login_required
def usuario_eliminar(request, user_id):
    from django.contrib.auth.models import User
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos para eliminar usuarios.')
        return redirect('dashboard')
        
    usuario = get_object_or_404(User, id=user_id)
    if usuario == request.user:
        messages.error(request, 'No puedes desactivar tu propia cuenta.')
        return redirect('usuarios_lista')
        
    if request.method == 'POST':
        try:
            username = usuario.username
            usuario.is_active = False
            usuario.save()
            
            registrar_auditoria(
                usuario=request.user,
                accion="Usuario Desactivado",
                registro_id=usuario.id,
                tabla_afectada="auth_user",
                estado_anterior="Activo",
                estado_nuevo="Inactivo",
                observacion=f"Usuario '{username}' desactivado (soft delete)"
            )
            messages.success(request, f'Usuario {username} desactivado correctamente.')
        except Exception as e:
            messages.error(request, f'Error al desactivar: {str(e)}')
            
    return redirect('usuarios_lista')

@login_required
def api_buscar_huesped(request):
    """API para buscar huésped registrado en la BD local por documento o nombre."""
    from django.http import JsonResponse
    from huespedes.models import Huesped
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 2:
        return JsonResponse({'results': []})
    from django.db.models import Q
    huespedes = Huesped.objects.filter(
        Q(num_doc__icontains=q) |
        Q(nombres__icontains=q) |
        Q(apellidos__icontains=q)
    ).order_by('apellidos', 'nombres')[:10]
    results = []
    for h in huespedes:
        results.append({
            'id': h.id,
            'num_doc': h.num_doc,
            'tipo_doc': h.tipo_doc,
            'nombres': h.nombres,
            'apellidos': h.apellidos,
            'nombre_completo': f'{h.nombres} {h.apellidos}',
            'telefono': h.telefono or '',
            'email': h.email or '',
            'nacionalidad': h.nacionalidad,
        })
    return JsonResponse({'results': results})


@login_required
def api_consulta_dni(request):
    from django.http import JsonResponse
    numero = request.GET.get('numero')
    if not numero or len(numero) != 8:
        return JsonResponse({'error': 'DNI inválido'}, status=400)
    
    try:
        import urllib.request, json
        req = urllib.request.Request(f'https://api.apis.net.pe/v1/dni?numero={numero}', headers={'User-Agent': 'Mozilla/5.0'})
        response = urllib.request.urlopen(req)
        data = json.loads(response.read().decode('utf-8'))
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def reserva_editar(request, reserva_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if reserva.estado in ['CHECKIN', 'CHECKOUT', 'CANCELADA']:
        messages.error(request, 'No se puede modificar una reserva en estado check-in, check-out o cancelada.')
        return redirect('reserva_detalle', reserva_id=reserva.id)

    if request.method == 'POST':
        try:
            huesped_id = request.POST.get('huesped_id')
            if huesped_id:
                reserva.huesped = Huesped.objects.get(id=huesped_id)
            else:
                reserva.huesped = Huesped.objects.get(num_doc=request.POST['huesped_doc'])

            reserva.habitacion = Habitacion.objects.get(id=request.POST['habitacion'])
            reserva.modalidad = request.POST.get('modalidad', Reserva.POR_DIA)
            reserva.num_adultos = int(request.POST.get('num_adultos', 1))
            reserva.origen = request.POST.get('origen', 'DIRECTO')
            reserva.observaciones = request.POST.get('observaciones', '')

            if reserva.modalidad == Reserva.POR_HORA:
                reserva.fecha_hora_entrada = parse_datetime_local(request.POST.get('fecha_hora_entrada'))
                if not reserva.fecha_hora_entrada:
                    raise ValidationError('Debes indicar la fecha y hora de ingreso.')
                reserva.duracion_horas = request.POST.get('duracion_horas') or 0
                reserva.fecha_entrada = reserva.fecha_hora_entrada.date()
                reserva.fecha_salida = (reserva.fecha_hora_entrada + timedelta(hours=float(reserva.duracion_horas or 3))).date()
            else:
                reserva.fecha_entrada = parse_date_local(request.POST.get('fecha_entrada'))
                reserva.fecha_salida = parse_date_local(
                    request.POST.get('fecha_salida') or request.POST.get('fecha_entrada')
                )
                if not reserva.fecha_entrada or not reserva.fecha_salida:
                    raise ValidationError('Debes indicar las fechas de entrada y salida.')
                reserva.fecha_hora_entrada = None
                reserva.duracion_horas = 0

            # Validar y salvar
            reserva.precio_total = reserva.calcular_precio()
            reserva.save()

            # Registrar auditoria
            # removed import, using utils.auditoria.log_action
            registrar_auditoria(
                usuario=request.user,
                accion="Modificar Reserva",
                registro_id=reserva.id,
                tabla_afectada="reservas_reserva",
                estado_nuevo=f"ID: {reserva.id}, Huesped: {reserva.huesped.nombres} {reserva.huesped.apellidos}, Hab: {reserva.habitacion.numero}, Total: S/. {reserva.precio_total}"
            )

            messages.success(request, f'Reserva #{reserva.id} modificada correctamente.')
            return redirect('reserva_detalle', reserva_id=reserva.id)
        except Exception as e:
            messages.error(request, f'Error al modificar la reserva: {str(e)}')

    selected_habitacion_id = reserva.habitacion.id if reserva.habitacion else ''
    selected_huesped = reserva.huesped

    return render(request, 'reservas/editar.html', {
        'reserva': reserva,
        'huespedes': Huesped.objects.all().order_by('nombres', 'apellidos'),
        'habitaciones': Habitacion.objects.exclude(estado='MANTENIMIENTO').select_related('tipo').order_by('piso', 'numero'),
        'selected_habitacion_id': selected_habitacion_id,
        'selected_huesped': selected_huesped,
        'hora_actual': timezone.localtime().strftime('%Y-%m-%dT%H:%M'),
    })


@login_required
def reserva_cancelar(request, reserva_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if reserva.estado in ['CHECKIN', 'CHECKOUT', 'CANCELADA']:
        messages.error(request, 'No se puede cancelar una reserva en estado check-in, check-out o cancelada.')
        return redirect('reserva_detalle', reserva_id=reserva.id)

    if request.method == 'POST':
        motivo = request.POST.get('motivo_cancelacion', '').strip()
        if not motivo:
            messages.error(request, 'Debes ingresar un motivo de cancelación.')
            return redirect('reserva_detalle', reserva_id=reserva.id)

        # Reembolsar anticipos automáticamente si existen
        pagos_anticipados = Pago.objects.filter(reserva=reserva, folio__isnull=True)
        total_reembolsado = Decimal('0.00')
        for pago in pagos_anticipados:
            from estancias.models import Reembolso
            Reembolso.objects.create(
                pago=pago,
                monto=pago.monto,
                motivo=f"Cancelación de reserva #{reserva.id}. {motivo}",
                estado=Reembolso.APROBADO,
                solicitado_por=request.user,
                aprobado_por=request.user,
                fecha_resolucion=timezone.now(),
                observacion="Reembolso automático por cancelación de reserva"
            )
            total_reembolsado += pago.monto

        if total_reembolsado > 0:
            reserva.estado = Reserva.REEMBOLSADO
            reserva.motivo_cancelacion = f"{motivo} (Reembolsado: S/. {total_reembolsado})"
        else:
            reserva.estado = Reserva.CANCELADA
            reserva.motivo_cancelacion = motivo
        reserva.save()

        if reserva.habitacion:
            reserva.habitacion.estado = Habitacion.DISPONIBLE
            reserva.habitacion.save()

        # Registrar auditoria
        # removed import, using utils.auditoria.log_action
        obs_extra = f" Reembolsado: S/. {total_reembolsado}" if total_reembolsado > 0 else ""
        registrar_auditoria(
            usuario=request.user,
            accion="Cancelar Reserva",
            registro_id=reserva.id,
            tabla_afectada="reservas_reserva",
            estado_nuevo=f"Estado: CANCELADA, Motivo: {motivo}{obs_extra}"
        )

        if total_reembolsado > 0:
            messages.success(request, f'Reserva #{reserva.id} cancelada. Se reembolsaron S/. {total_reembolsado} automáticamente.')
        else:
            messages.success(request, f'Reserva #{reserva.id} cancelada correctamente.')
        return redirect('reservas_lista')

    return redirect('reserva_detalle', reserva_id=reserva.id)


@login_required
def registrar_pago_anticipo(request, reserva_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if request.method == 'POST':
        monto = request.POST.get('monto')
        metodo_pago = request.POST.get('metodo_pago', Pago.EFECTIVO)
        transaccion_id = request.POST.get('transaccion_id') or None

        if monto:
            try:
                monto_decimal = Decimal(monto)
                if monto_decimal <= 0:
                    messages.error(request, 'El monto debe ser mayor a cero.')
                elif monto_decimal > reserva.saldo_pendiente:
                    messages.error(request, f'No se puede pagar más del saldo pendiente (S/. {reserva.saldo_pendiente:.2f}).')
                elif metodo_pago != Pago.EFECTIVO and not transaccion_id:
                    messages.error(request, 'Debes ingresar el ID de transacción para pagos electrónicos/bancarios.')
                else:
                    pago = Pago.objects.create(
                        reserva=reserva,
                        monto=monto_decimal,
                        metodo_pago=metodo_pago,
                        transaccion_id=transaccion_id
                    )
                    
                    # Registrar auditoria
                    # removed import, using utils.auditoria.log_action
                    registrar_auditoria(
                        usuario=request.user,
                        accion="Registrar Pago Anticipado",
                        registro_id=pago.id,
                        tabla_afectada="estancias_pago",
                        estado_nuevo=f"ID: {pago.id}, Reserva: {reserva.id}, Monto: S/. {monto_decimal}, Metodo: {metodo_pago}"
                    )
                    
                    messages.success(request, 'Pago anticipado registrado correctamente.')
            except Exception as e:
                messages.error(request, f'Error al registrar pago: {str(e)}')
        else:
            messages.error(request, 'Monto inválido para el pago.')
            
    return redirect('reserva_detalle', reserva_id=reserva_id)


@login_required
def solicitar_reembolso(request, pago_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    pago = get_object_or_404(Pago, id=pago_id)
    
    redirect_url = redirect('dashboard')
    if pago.folio:
        redirect_url = redirect('folio', estancia_id=pago.folio.estancia.id)
    elif pago.reserva:
        redirect_url = redirect('reserva_detalle', reserva_id=pago.reserva.id)

    if request.method == 'POST':
        from django.db import transaction
        monto_str = request.POST.get('monto')
        motivo = request.POST.get('motivo', '').strip()
        cancelar_estancia_problema = 'cancelar_estancia_problema' in request.POST
        estado_habitacion = request.POST.get('estado_habitacion', 'MANTENIMIENTO')
        
        if not monto_str or not motivo:
            messages.error(request, 'Debes ingresar monto y motivo del reembolso.')
            return redirect_url

        try:
            monto_decimal = Decimal(monto_str)
            if monto_decimal <= 0:
                messages.error(request, 'El monto del reembolso debe ser mayor a cero.')
                return redirect_url

            from estancias.models import Reembolso
            reembolsos_previos = Reembolso.objects.filter(
                pago=pago,
                estado__in=[Reembolso.SOLICITADO, Reembolso.APROBADO]
            )
            total_reembolsado = sum(r.monto for r in reembolsos_previos)
            monto_disponible = pago.monto - total_reembolsado
            
            if monto_disponible <= 0:
                messages.error(request, 'Este pago ya fue completamente reembolsado.')
                return redirect_url
                
            if monto_decimal > monto_disponible:
                messages.error(request, f'Monto disponible para reembolsar: S/. {monto_disponible:.2f}')
                return redirect_url
                
            with transaction.atomic():
                if cancelar_estancia_problema:
                    # Caso 2: Cancelación total por problemas de la habitación (Responsabilidad del hotel)
                    estancia = pago.folio.estancia if pago.folio else getattr(pago.reserva, 'estancia', None)
                    reserva = pago.reserva if pago.reserva else (estancia.reserva if estancia else None)
                    
                    if estancia and estancia.estado == Estancia.ACTIVA:
                        # 1. Exonerar todos los cargos de la estancia
                        cargos = CargoEstancia.objects.filter(estancia=estancia)
                        for cargo in cargos:
                            cargo.exonerado = True
                            cargo.motivo_exoneracion = f"Cancelación por problema en la habitación: {motivo}"
                            cargo.exonerado_por = request.user
                            cargo.save()
                            
                            # Auditoría de cada cargo exonerado
                            # removed import, using utils.auditoria.log_action
                            registrar_auditoria(
                                usuario=request.user,
                                accion="Cargo Exonerado (Problema de Hab.)",
                                registro_id=cargo.id,
                                tabla_afectada="estancias_cargoestancia",
                                estado_nuevo=f"ID: {cargo.id}, Concepto: {cargo.concepto}, Monto: S/. {cargo.monto} -> Exonerado por problema"
                            )
                        
                        # 2. Recalcular totales del folio
                        folio = estancia.folio
                        folio.calcular_totales()
                        
                        # 3. Finalizar la estancia con precio_final 0 y registrar check-out
                        estancia.estado = Estancia.FINALIZADA
                        estancia.fecha_checkout = timezone.now()
                        estancia.precio_final = Decimal('0.00')
                        estancia.save()
                        
                        # 4. Cerrar Folio
                        folio.estado = Folio.CERRADO
                        folio.save()
                        
                        # 5. Cancelar Reserva (se guarda antes de poner la habitación en mantenimiento para evitar ValidationError en clean())
                        if reserva:
                            reserva.estado = Reserva.CANCELADA
                            reserva.motivo_cancelacion = f"Cancelada por problema en habitación ({estancia.habitacion.numero}): {motivo}"
                            reserva.save()
                            
                        # 6. Cambiar estado de la habitación (Mantenimiento o Limpieza)
                        habitacion = estancia.habitacion
                        habitacion.estado = estado_habitacion
                        habitacion.save()
                            
                        # Registrar Auditoría de Cancelación de Estancia
                        # removed import, using utils.auditoria.log_action
                        registrar_auditoria(
                            usuario=request.user,
                            accion="Cancelar Estancia (Problema de Hab.)",
                            registro_id=estancia.id,
                            tabla_afectada="estancias_estancia",
                            estado_nuevo=f"ID: {estancia.id}, Hab: {habitacion.numero} -> {estado_habitacion}, Reserva: {reserva.id if reserva else 'N/A'}"
                        )
                    elif reserva and reserva.estado in [Reserva.PENDIENTE, Reserva.CONFIRMADA]:
                        # Cancelar reserva antes del check-in
                        reserva.estado = Reserva.CANCELADA
                        reserva.motivo_cancelacion = f"Cancelada por problema en habitación antes de check-in: {motivo}"
                        reserva.save()
                        if reserva.habitacion:
                            reserva.habitacion.estado = Habitacion.DISPONIBLE
                            reserva.habitacion.save()
                    
                    # 7. Crear el reembolso directamente APROBADO por ser un problema de habitación
                    reembolso = Reembolso.objects.create(
                        pago=pago,
                        monto=monto_decimal,
                        motivo=motivo,
                        estado=Reembolso.APROBADO,
                        solicitado_por=request.user,
                        aprobado_por=request.user,
                        fecha_resolucion=timezone.now(),
                        observacion=f"Aprobado automáticamente por cancelación de estancia debido a problemas de habitación: {motivo}"
                    )
                    
                    # Registrar auditoría de reembolso aprobado
                    # removed import, using utils.auditoria.log_action
                    registrar_auditoria(
                        usuario=request.user,
                        accion="Resolver Reembolso",
                        registro_id=reembolso.id,
                        tabla_afectada="estancias_reembolso",
                        estado_nuevo=f"ID: {reembolso.id}, Estado: APROBADO (Aut. por Problema Hab.), Monto: S/. {monto_decimal}"
                    )
                    
                    messages.success(request, 'Estancia cancelada, cargos exonerados, habitación liberada y reembolso procesado correctamente.')
                else:
                    # Caso 1: Reembolso normal (comportamiento por defecto)
                    reembolso = Reembolso.objects.create(
                        pago=pago,
                        monto=monto_decimal,
                        motivo=motivo,
                        estado=Reembolso.SOLICITADO,
                        solicitado_por=request.user
                    )
                    
                    # Registrar auditoria
                    # removed import, using utils.auditoria.log_action
                    registrar_auditoria(
                        usuario=request.user,
                        accion="Solicitar Reembolso",
                        registro_id=reembolso.id,
                        tabla_afectada="estancias_reembolso",
                        estado_nuevo=f"ID: {reembolso.id}, Pago: {pago.id}, Monto: S/. {monto_decimal}"
                    )
                    
                    messages.success(request, 'Solicitud de reembolso registrada correctamente.')
        except Exception as e:
            messages.error(request, f'Error al registrar la solicitud: {str(e)}')
            
    return redirect_url


@login_required
def cancelar_estancia_sin_pago(request, estancia_id):
    """
    Obs. 1 – Escenario 3: Cancela una estancia activa cuando NO existe pago anticipado.
    No genera ningún objeto Reembolso. Exonera cargos, finaliza estancia/reserva
    y cambia estado de la habitación a Mantenimiento o Limpieza.
    """
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    estancia = get_object_or_404(Estancia, id=estancia_id)

    if request.method == 'POST':
        motivo = request.POST.get('motivo', '').strip()
        estado_habitacion = request.POST.get('estado_habitacion_nopago', 'MANTENIMIENTO')

        if not motivo:
            messages.error(request, 'El motivo de la cancelación es obligatorio.')
            return redirect('folio', estancia_id=estancia_id)

        # Guardia: verificar que realmente no hay pagos registrados
        folio = getattr(estancia, 'folio', None)
        if folio and folio.total_pagado > 0:
            messages.error(
                request,
                'Esta estancia tiene pagos registrados. Use la opción "Solicitar Reembolso" para cancelarla.'
            )
            return redirect('folio', estancia_id=estancia_id)

        if estancia.estado != Estancia.ACTIVA:
            messages.error(request, 'Solo se pueden cancelar estancias activas.')
            return redirect('folio', estancia_id=estancia_id)

        try:
            from django.db import transaction
            # removed import, using utils.auditoria.log_action
            with transaction.atomic():
                # 1. Exonerar todos los cargos del folio
                cargos = CargoEstancia.objects.filter(estancia=estancia)
                for cargo in cargos:
                    cargo.exonerado = True
                    cargo.motivo_exoneracion = f"Cancelación por problema de habitación (sin pago previo): {motivo}"
                    cargo.exonerado_por = request.user
                    cargo.save()
                    registrar_auditoria(
                        usuario=request.user,
                        accion="Cargo Exonerado (Cancelación Sin Pago)",
                        registro_id=cargo.id,
                        tabla_afectada="estancias_cargoestancia",
                        estado_nuevo=f"ID: {cargo.id}, {cargo.concepto} -> Exonerado (sin pago)"
                    )

                # 2. Recalcular totales del folio -> saldo queda en S/. 0.00
                if folio:
                    folio.calcular_totales()
                    folio.estado = Folio.CERRADO
                    folio.save()

                # 3. Finalizar estancia
                estancia.estado = Estancia.FINALIZADA
                estancia.fecha_checkout = timezone.now()
                estancia.precio_final = Decimal('0.00')
                estancia.save()

                # 4. Cancelar reserva asociada
                reserva = estancia.reserva
                if reserva:
                    reserva.estado = Reserva.CANCELADA
                    reserva.motivo_cancelacion = (
                        f"Cancelada por problema de habitación "
                        f"({estancia.habitacion.numero}) sin pago previo: {motivo}"
                    )
                    reserva.save()

                # 5. Cambiar estado de la habitación
                habitacion = estancia.habitacion
                habitacion.estado = estado_habitacion
                habitacion.save()

                # 6. Auditoría de cancelación sin pago
                registrar_auditoria(
                    usuario=request.user,
                    accion="Cancelar Estancia Sin Pago (Problema Hab.)",
                    registro_id=estancia.id,
                    tabla_afectada="estancias_estancia",
                    estado_nuevo=(
                        f"ID: {estancia.id}, Hab: {habitacion.numero} -> {estado_habitacion}, "
                        f"Reserva: {reserva.id if reserva else 'N/A'}, Sin reembolso"
                    )
                )

            messages.success(
                request,
                f'Estancia cancelada correctamente. Cargos exonerados. '
                f'Habitación {habitacion.numero} enviada a {habitacion.get_estado_display()}. '
                f'No se generó reembolso ya que no existía pago registrado.'
            )
        except Exception as e:
            messages.error(request, f'Error al cancelar la estancia: {str(e)}')

    return redirect('dashboard')


@login_required
def cambiar_habitacion(request, estancia_id):

    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
    estancia = get_object_or_404(Estancia, id=estancia_id)
    
    if request.method == 'POST':
        nueva_hab_id = request.POST.get('nueva_habitacion_id')
        motivo = request.POST.get('motivo', '').strip()
        
        if not nueva_hab_id or not motivo:
            messages.error(request, 'Nueva habitación y motivo son requeridos.')
            return redirect('folio', estancia_id=estancia_id)
        
        try:
            import estancias.services as estancia_services
            estancia_services.cambiar_habitacion_activo(
                estancia_id=estancia_id,
                nueva_habitacion_id=nueva_hab_id,
                motivo=motivo,
                usuario=request.user
            )
            messages.success(request, 'Cambio de habitación completado correctamente.')
        except ValidationError as e:
            messages.error(request, e.message if hasattr(e, 'message') else str(e))
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return redirect('folio', estancia_id=estancia_id)


@login_required
def aprobar_reembolso(request, reembolso_id):
    if not _es_admin(request.user):
        return _acceso_denegado(request, 'Solo los administradores pueden resolver reembolsos.')
        
    from estancias.models import Reembolso
    reembolso = get_object_or_404(Reembolso, id=reembolso_id)
    
    redirect_url = redirect('dashboard')
    if reembolso.pago.folio:
        redirect_url = redirect('folio', estancia_id=reembolso.pago.folio.estancia.id)
    elif reembolso.pago.reserva:
        redirect_url = redirect('reserva_detalle', reserva_id=reembolso.pago.reserva.id)

    if request.method == 'POST':
        accion = request.POST.get('accion')
        observacion = request.POST.get('observacion', '').strip()
        
        if reembolso.estado != Reembolso.SOLICITADO:
            messages.error(request, 'Este reembolso ya ha sido resuelto.')
            return redirect_url

        if accion in ['APROBAR', 'aprobar']:
            reembolso.estado = Reembolso.APROBADO
            reembolso.aprobado_por = request.user
            reembolso.fecha_resolucion = timezone.now()
            reembolso.observacion = observacion
            reembolso.save()

            pago = reembolso.pago
            estancia = pago.folio.estancia if pago.folio else None
            reserva = pago.reserva if pago.reserva else (estancia.reserva if estancia else None)
            
            if estancia and estancia.estado == Estancia.ACTIVA:
                folio = estancia.folio
                folio.calcular_totales()
                
                # Cerrar folio y perdonar saldo pendiente
                folio.estado = Folio.CERRADO
                folio.save()
                
                estancia.habitacion.estado = Habitacion.LIMPIEZA
                estancia.habitacion.save()
                estancia.fecha_checkout = timezone.now()
                estancia.estado = Estancia.FINALIZADA
                estancia.save()
                
                if reserva:
                    reserva.estado = Reserva.REEMBOLSADO
                    reserva.motivo_cancelacion = f"Reembolso #{reembolso.id} aprobado. {observacion}"
                    reserva.save()
                
                messages.success(request, f'Reembolso aprobado. Estancia finalizada, habitación en limpieza y reserva reembolsada.')
            elif reserva and reserva.estado in [Reserva.PENDIENTE, Reserva.CONFIRMADA]:
                if reserva.habitacion:
                    reserva.habitacion.estado = Habitacion.DISPONIBLE
                    reserva.habitacion.save()
                reserva.estado = Reserva.REEMBOLSADO
                reserva.motivo_cancelacion = f"Reembolso anticipo #{reembolso.id} aprobado. {observacion}"
                reserva.save()
                messages.success(request, f'Reembolso aprobado. Reserva reembolsada y habitación liberada.')
            else:
                messages.success(request, f'Reembolso #{reembolso.id} aprobado correctamente.')
        elif accion in ['RECHAZAR', 'rechazar']:
            reembolso.estado = Reembolso.RECHAZADO
        else:
            messages.error(request, 'Acción no válida.')
            return redirect_url
            
        reembolso.aprobado_por = request.user
        reembolso.fecha_resolucion = timezone.now()
        reembolso.observacion = observacion
        reembolso.save()
        
        # Registrar auditoria
        # removed import, using utils.auditoria.log_action
        registrar_auditoria(
            usuario=request.user,
            accion="Resolver Reembolso",
            registro_id=reembolso.id,
            tabla_afectada="estancias_reembolso",
            estado_nuevo=f"ID: {reembolso.id}, Estado: {reembolso.estado}, Observacion: {observacion}"
        )
        
        messages.success(request, f'Reembolso #{reembolso.id} resuelto: {reembolso.get_estado_display()}.')
        
    return redirect_url


# ==============================================================================
#  ÉPICO 05 – ATENCIÓN AL CLIENTE (TICKETS DE SERVICIO) Y REEMBOLSOS
# ==============================================================================
from django.db import transaction
from atencion.models import TicketServicio, SeguimientoTicket

@login_required
def tickets_lista(request):
    if not _es_recepcionista(request.user) and not _es_housekeeping(request.user):
        return _acceso_denegado(request)
        
    tickets = TicketServicio.objects.select_related('estancia__habitacion', 'estancia__reserva__huesped', 'recepcionista').all()
    
    # Filtros
    cliente_q = request.GET.get('cliente', '').strip()
    habitacion_q = request.GET.get('habitacion', '').strip()
    estado_q = request.GET.get('estado', '').strip()
    categoria_q = request.GET.get('categoria', '').strip()
    prioridad_q = request.GET.get('prioridad', '').strip()
    responsable_q = request.GET.get('responsable', '').strip()
    fecha_q = request.GET.get('fecha', '').strip()

    if cliente_q:
        tickets = tickets.filter(
            Q(estancia__reserva__huesped__nombres__icontains=cliente_q) |
            Q(estancia__reserva__huesped__apellidos__icontains=cliente_q) |
            Q(estancia__reserva__huesped__num_doc__icontains=cliente_q)
        )
    if habitacion_q:
        tickets = tickets.filter(estancia__habitacion__numero=habitacion_q)
    if estado_q:
        tickets = tickets.filter(estado=estado_q)
    if categoria_q:
        tickets = tickets.filter(categoria=categoria_q)
    if prioridad_q:
        tickets = tickets.filter(prioridad=prioridad_q)
    if responsable_q:
        tickets = tickets.filter(responsable=responsable_q)
    if fecha_q:
        try:
            fecha_d = date.fromisoformat(fecha_q)
            tickets = tickets.filter(fecha=fecha_d)
        except ValueError:
            pass

    return render(request, 'atencion/lista.html', {
        'tickets': tickets,
        'categorias': TicketServicio.CATEGORIAS,
        'prioridades': TicketServicio.PRIORIDADES,
        'estados': TicketServicio.ESTADOS,
        'areas': TicketServicio.AREAS,
        'f_cliente': cliente_q,
        'f_habitacion': habitacion_q,
        'f_estado': estado_q,
        'f_categoria': categoria_q,
        'f_prioridad': prioridad_q,
        'f_responsable': responsable_q,
        'f_fecha': fecha_q,
    })


@login_required
def ticket_nuevo(request):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
        
    estancias_activas = Estancia.objects.filter(estado=Estancia.ACTIVA).select_related('habitacion', 'reserva__huesped')
    
    if request.method == 'POST':
        estancia_id = request.POST.get('estancia')
        categoria = request.POST.get('categoria')
        prioridad = request.POST.get('prioridad', 'MEDIA')
        responsable = request.POST.get('responsable', 'RECEPCION')
        descripcion = request.POST.get('descripcion', '').strip()
        
        if not estancia_id or not categoria or not descripcion:
            messages.error(request, 'Todos los campos marcados como obligatorios son requeridos.')
            return redirect('ticket_nuevo')
            
        try:
            estancia = Estancia.objects.get(id=estancia_id)
            if estancia.estado != Estancia.ACTIVA:
                messages.error(request, 'La habitación seleccionada no tiene un hospedaje activo.')
                return redirect('ticket_nuevo')
                
            with transaction.atomic():
                ticket = TicketServicio.objects.create(
                    estancia=estancia,
                    categoria=categoria,
                    prioridad=prioridad,
                    responsable=responsable,
                    descripcion=descripcion,
                    recepcionista=request.user
                )
                
                # Crear seguimiento inicial
                SeguimientoTicket.objects.create(
                    ticket=ticket,
                    usuario=request.user,
                    comentario=f"Se abre el ticket de atención en estado Abierta. Descripción inicial: {descripcion}",
                    estado_ticket=ticket.estado
                )
                
                # Automatizaciones (Limpieza -> Housekeeping)
                if categoria == 'LIMPIEZA':
                    habitacion = estancia.habitacion
                    habitacion.estado = Habitacion.LIMPIEZA
                    habitacion.save()
                    
                    # Registrar auditoría de housekeeping automático
                    # removed import, using utils.auditoria.log_action
                    registrar_auditoria(
                        usuario=request.user,
                        accion="Housekeeping Estado Modificado (Aut.)",
                        registro_id=habitacion.id,
                        tabla_afectada="hotel_habitacion",
                        estado_nuevo=Habitacion.LIMPIEZA,
                        observacion=f"Ticket {ticket.numero_atencion} de limpieza abrió orden automática para Hab. {habitacion.numero}"
                    )
                elif categoria == 'MANTENIMIENTO':
                    habitacion = estancia.habitacion
                    habitacion.estado = Habitacion.MANTENIMIENTO
                    habitacion.save()
                    
                    # Registrar auditoría de mantenimiento automático
                    # removed import, using utils.auditoria.log_action
                    registrar_auditoria(
                        usuario=request.user,
                        accion="Habitación puesta en Mantenimiento (Aut.)",
                        registro_id=habitacion.id,
                        tabla_afectada="hotel_habitacion",
                        estado_nuevo=Habitacion.MANTENIMIENTO,
                        observacion=f"Ticket {ticket.numero_atencion} de mantenimiento bloqueó Hab. {habitacion.numero}"
                    )
                
                # Registrar auditoría del ticket creado
                # removed import, using utils.auditoria.log_action
                registrar_auditoria(
                    usuario=request.user,
                    accion="Crear Ticket de Servicio",
                    registro_id=ticket.id,
                    tabla_afectada="atencion_ticketservicio",
                    estado_nuevo=f"Ticket {ticket.numero_atencion} creado para Hab. {estancia.habitacion.numero}"
                )
                
                messages.success(request, f'Ticket {ticket.numero_atencion} registrado correctamente.')
                return redirect('ticket_detalle', ticket_id=ticket.id)
        except Exception as e:
            messages.error(request, f'Error al registrar el ticket: {str(e)}')
            
    return render(request, 'atencion/nueva.html', {
        'estancias': estancias_activas,
        'categorias': TicketServicio.CATEGORIAS,
        'prioridades': TicketServicio.PRIORIDADES,
        'areas': TicketServicio.AREAS,
    })


@login_required
def ticket_detalle(request, ticket_id):
    if not _es_recepcionista(request.user) and not _es_housekeeping(request.user):
        return _acceso_denegado(request)
        
    ticket = get_object_or_404(TicketServicio.objects.select_related('estancia__habitacion', 'estancia__reserva__huesped', 'recepcionista'), id=ticket_id)
    seguimientos = ticket.seguimientos.select_related('usuario').all()
    
    return render(request, 'atencion/detalle.html', {
        'ticket': ticket,
        'seguimientos': seguimientos,
        'estados': TicketServicio.ESTADOS,
        'areas': TicketServicio.AREAS,
    })


@login_required
def ticket_iniciar(request, ticket_id):
    if not _es_recepcionista(request.user) and not _es_housekeeping(request.user):
        return _acceso_denegado(request)
        
    ticket = get_object_or_404(TicketServicio, id=ticket_id)
    if request.method == 'POST':
        if ticket.estado != 'ABIERTA':
            messages.error(request, 'El ticket ya fue iniciado o resuelto.')
            return redirect('ticket_detalle', ticket_id=ticket.id)
            
        with transaction.atomic():
            ticket.estado = 'PROCESO'
            ticket.save()
            
            SeguimientoTicket.objects.create(
                ticket=ticket,
                usuario=request.user,
                comentario="Se inicia el trabajo sobre la solicitud. Estado cambiado a En Proceso.",
                estado_ticket=ticket.estado
            )
            messages.success(request, 'Trabajo iniciado.')
            
    return redirect('ticket_detalle', ticket_id=ticket.id)


@login_required
def ticket_resolver(request, ticket_id):
    if not _es_recepcionista(request.user) and not _es_housekeeping(request.user):
        return _acceso_denegado(request)
        
    ticket = get_object_or_404(TicketServicio, id=ticket_id)
    if request.method == 'POST':
        solucion = request.POST.get('solucion', '').strip()
        observacion = request.POST.get('observacion_resolucion', '').strip()
        
        if not solucion:
            messages.error(request, 'Debes detallar la solución implementada.')
            return redirect('ticket_detalle', ticket_id=ticket.id)
            
        with transaction.atomic():
            ticket.estado = 'RESUELTA'
            ticket.solucion = solucion
            ticket.observacion_resolucion = observacion
            ticket.resolved_at = timezone.now()
            ticket.save()
            
            SeguimientoTicket.objects.create(
                ticket=ticket,
                usuario=request.user,
                comentario=f"Solicitud resuelta. Solución: {solucion}. Observaciones: {observacion}",
                estado_ticket=ticket.estado
            )
            
            # Automatización: Si el estado de la habitación estaba en LIMPIEZA y se resuelve, vuelve a OCUPADA
            if ticket.categoria == 'LIMPIEZA' and ticket.estancia.habitacion.estado == Habitacion.LIMPIEZA:
                habitacion = ticket.estancia.habitacion
                habitacion.estado = Habitacion.OCUPADA if ticket.estancia.estado == Estancia.ACTIVA else Habitacion.DISPONIBLE
                habitacion.save()
                
                # Registrar auditoría de housekeeping automático resuelto
                # removed import, using utils.auditoria.log_action
                registrar_auditoria(
                    usuario=request.user,
                    accion="Housekeeping Completado (Ticket)",
                    registro_id=habitacion.id,
                    tabla_afectada="hotel_habitacion",
                    estado_nuevo=habitacion.estado,
                    observacion=f"Habitación {habitacion.numero} marcada como {habitacion.estado} al resolverse ticket {ticket.numero_atencion}"
                )
                
            messages.success(request, 'Solicitud marcada como Resuelta.')
            
    return redirect('ticket_detalle', ticket_id=ticket.id)


@login_required
def ticket_cerrar(request, ticket_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
        
    ticket = get_object_or_404(TicketServicio, id=ticket_id)
    if request.method == 'POST':
        if ticket.estado != 'RESUELTA':
            messages.error(request, 'Solo se pueden cerrar tickets previamente resueltos.')
            return redirect('ticket_detalle', ticket_id=ticket.id)
            
        with transaction.atomic():
            ticket.estado = 'CERRADA'
            ticket.closed_at = timezone.now()
            ticket.save()
            
            SeguimientoTicket.objects.create(
                ticket=ticket,
                usuario=request.user,
                comentario="Atención cerrada definitivamente.",
                estado_ticket=ticket.estado
            )
            messages.success(request, 'Ticket cerrado definitivamente.')
            
    return redirect('ticket_detalle', ticket_id=ticket.id)


@login_required
def ticket_reabrir(request, ticket_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
        
    ticket = get_object_or_404(TicketServicio, id=ticket_id)
    if request.method == 'POST':
        motivo = request.POST.get('motivo_reapertura', '').strip()
        if not motivo:
            messages.error(request, 'Debes indicar el motivo de la reapertura.')
            return redirect('ticket_detalle', ticket_id=ticket.id)
            
        with transaction.atomic():
            ticket.estado = 'PROCESO'  # Reabierta -> pasa a En Proceso
            ticket.motivo_reapertura = motivo
            ticket.solucion = None
            ticket.observacion_resolucion = None
            ticket.resolved_at = None
            ticket.closed_at = None
            ticket.save()
            
            SeguimientoTicket.objects.create(
                ticket=ticket,
                usuario=request.user,
                comentario=f"Ticket reabierto. Motivo: {motivo}",
                estado_ticket=ticket.estado
            )
            messages.success(request, 'Ticket reabierto y asignado en proceso.')
            
    return redirect('ticket_detalle', ticket_id=ticket.id)


@login_required
def ticket_seguimiento(request, ticket_id):
    if not _es_recepcionista(request.user) and not _es_housekeeping(request.user):
        return _acceso_denegado(request)
        
    ticket = get_object_or_404(TicketServicio, id=ticket_id)
    if request.method == 'POST':
        comentario = request.POST.get('comentario', '').strip()
        nuevo_estado = request.POST.get('estado')
        
        if not comentario:
            messages.error(request, 'Debes ingresar un comentario.')
            return redirect('ticket_detalle', ticket_id=ticket.id)
            
        with transaction.atomic():
            if nuevo_estado and nuevo_estado in dict(TicketServicio.ESTADOS):
                ticket.estado = nuevo_estado
                ticket.save()
                
            SeguimientoTicket.objects.create(
                ticket=ticket,
                usuario=request.user,
                comentario=comentario,
                estado_ticket=ticket.estado
            )
            messages.success(request, 'Seguimiento registrado.')
            
    return redirect('ticket_detalle', ticket_id=ticket.id)


@login_required
def ticket_agregar_cargo(request, ticket_id):
    if not _es_recepcionista(request.user):
        return _acceso_denegado(request)
        
    ticket = get_object_or_404(TicketServicio, id=ticket_id)
    if request.method == 'POST':
        concepto = request.POST.get('concepto', '').strip()
        monto = request.POST.get('monto')
        tipo_cargo = request.POST.get('tipo', CargoEstancia.OTRO)
        
        if not concepto or not monto:
            messages.error(request, 'Concepto y monto son obligatorios.')
            return redirect('ticket_detalle', ticket_id=ticket.id)
            
        try:
            monto_decimal = Decimal(monto)
            if monto_decimal <= 0:
                messages.error(request, 'El monto debe ser mayor a cero.')
                return redirect('ticket_detalle', ticket_id=ticket.id)
                
            estancia = ticket.estancia
            with transaction.atomic():
                cargo = CargoEstancia.objects.create(
                    estancia=estancia,
                    concepto=f"{concepto} (Ticket: {ticket.numero_atencion})",
                    monto=monto_decimal,
                    tipo=tipo_cargo
                )
                
                # Recalcular totales del folio
                if hasattr(estancia, 'folio'):
                    estancia.folio.calcular_totales()
                    
                # Registrar seguimiento en el ticket
                SeguimientoTicket.objects.create(
                    ticket=ticket,
                    usuario=request.user,
                    comentario=f"Se asoció cargo extra al folio del hospedaje: {concepto} por S/. {monto_decimal}",
                    estado_ticket=ticket.estado
                )
                
                # Registrar auditoría de caja
                # removed import, using utils.auditoria.log_action
                registrar_auditoria(
                    usuario=request.user,
                    accion="Cargo Extra de Ticket Registrado",
                    registro_id=cargo.id,
                    tabla_afectada="estancias_cargoestancia",
                    estado_nuevo=f"ID: {cargo.id}, Ticket: {ticket.numero_atencion}, Monto: S/. {monto_decimal}"
                )
                
                messages.success(request, 'Cargo asociado al folio de la estancia correctamente.')
        except Exception as e:
            messages.error(request, f'Error al registrar cargo: {str(e)}')
            
    return redirect('ticket_detalle', ticket_id=ticket.id)


@login_required
def reembolsos_lista_admin(request):
    if not _es_admin(request.user):
        return _acceso_denegado(request, 'Solo los administradores pueden ver la lista de reembolsos.')
        
    from estancias.models import Reembolso
    reembolsos = Reembolso.objects.select_related('pago__folio__estancia__habitacion', 'pago__reserva__huesped', 'solicitado_por', 'aprobado_por').all().order_by('-fecha_solicitud')
    
    estado_q = request.GET.get('estado', '').strip()
    if estado_q:
        reembolsos = reembolsos.filter(estado=estado_q)
        
    return render(request, 'reportes/reembolsos.html', {
        'reembolsos': reembolsos,
        'f_estado': estado_q,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Obs. 2 – API de Ocupación en Tiempo Real y Alertas de Check-Out
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def api_ocupacion_habitaciones(request):
    """
    Obs. 2: Retorna JSON con el estado de ocupación en tiempo real de todas las
    estancias activas. Incluye tiempos transcurridos, restantes y el flag
    `proxima_salida` que activa la alerta visual de check-out próximo.
    """
    from django.http import JsonResponse
    from datetime import timedelta

    now = timezone.localtime(timezone.now())
    try:
        from hotel.models import Hotel
        hotel = Hotel.objects.first()
        alerta_minutos = getattr(hotel, 'alerta_checkout_minutos', 30) if hotel else 30
    except Exception:
        alerta_minutos = 30

    estancias_activas = Estancia.objects.filter(
        estado=Estancia.ACTIVA
    ).select_related('reserva', 'reserva__huesped', 'habitacion')

    data = []
    for est in estancias_activas:
        checkin_local = timezone.localtime(est.fecha_checkin)
        tiempo_transcurrido = now - checkin_local
        minutos_transcurridos = int(tiempo_transcurrido.total_seconds() // 60)
        horas_t = minutos_transcurridos // 60
        mins_t = minutos_transcurridos % 60

        fecha_salida = getattr(est.reserva, 'fecha_hora_salida', None)
        minutos_restantes = None
        proxima_salida = False
        salida_vencida = False
        salida_str = '—'

        if fecha_salida:
            salida_local = timezone.localtime(fecha_salida)
            salida_str = salida_local.strftime('%H:%M %d/%m')
            diff = salida_local - now
            minutos_restantes = int(diff.total_seconds() // 60)
            proxima_salida = 0 < minutos_restantes <= alerta_minutos
            salida_vencida = minutos_restantes < 0

        huesped = est.reserva.huesped
        data.append({
            'estancia_id': est.id,
            'habitacion': est.habitacion.numero,
            'huesped': f"{huesped.nombres} {huesped.apellidos}",
            'checkin': checkin_local.strftime('%H:%M'),
            'salida_estimada': salida_str,
            'minutos_transcurridos': minutos_transcurridos,
            'tiempo_transcurrido_str': f"{horas_t}h {mins_t:02d}m",
            'minutos_restantes': minutos_restantes,
            'proxima_salida': proxima_salida,
            'salida_vencida': salida_vencida,
            'folio_url': f"/estancias/{est.id}/folio/",
        })

    return JsonResponse({
        'estancias': data,
        'alerta_minutos': alerta_minutos,
        'total': len(data),
    })


@login_required
def api_habitaciones_housekeeping_recientes(request):
    """
    Obs. 2: Retorna habitaciones cuyo check-out se realizó hace ≤ 10 minutos.
    Estas deben ser inspeccionadas urgentemente por Housekeeping.
    """
    from django.http import JsonResponse
    from datetime import timedelta

    now = timezone.now()
    limite = now - timedelta(minutes=10)

    estancias_recientes = Estancia.objects.filter(
        estado=Estancia.FINALIZADA,
        habitacion__estado=Habitacion.LIMPIEZA,
        fecha_checkout__gte=limite
    ).select_related('habitacion', 'habitacion__tipo', 'reserva__huesped').order_by('-fecha_checkout')

    data = []
    for e in estancias_recientes:
        checkout_local = timezone.localtime(e.fecha_checkout)
        minutos_desde_checkout = int((now - e.fecha_checkout).total_seconds() // 60)
        huesped = e.reserva.huesped
        data.append({
            'estancia_id': e.id,
            'habitacion': e.habitacion.numero,
            'piso': e.habitacion.piso,
            'tipo': e.habitacion.tipo.nombre,
            'huesped': f"{huesped.nombres} {huesped.apellidos}",
            'checkout_hora': checkout_local.strftime('%H:%M'),
            'minutos_desde_checkout': minutos_desde_checkout,
            'urgente': True,
        })

    return JsonResponse({
        'habitaciones_urgentes': data,
        'total': len(data),
    })


@login_required
def mi_perfil(request):
    from django.contrib.auth.models import Group
    # import removed; using utils.auditoria.log_action
    
    if request.method == 'POST':
        try:
            request.user.first_name = request.POST.get('nombres', request.user.first_name)
            request.user.last_name = request.POST.get('apellidos', request.user.last_name)
            request.user.email = request.POST.get('email', request.user.email)
            request.user.save()
            
            log_action(
                user=request.user,
                accion="Perfil Actualizado",
                registro_id=request.user.id,
                tabla_afectada="auth_user",
                estado_nuevo="Datos de perfil actualizados"
            )
            
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('mi_perfil')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
    
    roles = list(request.user.groups.values_list('name', flat=True))
    
    return render(request, 'mi_perfil.html', {
        'user': request.user,
        'roles': roles,
    })


@login_required
def cambiar_password(request):
    if request.method == 'POST':
        password_actual = request.POST.get('password_actual')
        nueva_password = request.POST.get('nueva_password')
        confirmar_password = request.POST.get('confirmar_password')
        
        if not request.user.check_password(password_actual):
            messages.error(request, 'La contraseña actual es incorrecta.')
            return redirect('cambiar_password')
        
        if nueva_password != confirmar_password:
            messages.error(request, 'Las contraseñas nuevas no coinciden.')
            return redirect('cambiar_password')
        
        if len(nueva_password) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return redirect('cambiar_password')
        
        if not any(c.isupper() for c in nueva_password):
            messages.error(request, 'La contraseña debe contener al menos una letra mayúscula.')
            return redirect('cambiar_password')
        
        if not any(c.isdigit() for c in nueva_password):
            messages.error(request, 'La contraseña debe contener al menos un número.')
            return redirect('cambiar_password')
        
        request.user.set_password(nueva_password)
        request.user.save()
        
        registrar_auditoria(
            usuario=request.user,
            accion="Contraseña Cambiada",
            registro_id=request.user.id,
            tabla_afectada="auth_user",
            estado_nuevo="Contraseña actualizada exitosamente"
        )
        
        messages.success(request, 'Contraseña cambiada correctamente. Inicia sesión nuevamente.')
        return redirect('login')
    
    return render(request, 'cambiar_password.html')


@login_required
def recuperar_contrasena(request):
    """Vista para solicitar recuperación de contraseña"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            messages.error(request, 'Debes ingresar un correo electrónico.')
            return redirect('recuperar_contrasena')
        
        from django.contrib.auth.models import User
        from reportes.models import PasswordResetToken
        from django.core.mail import send_mail
        from django.conf import settings
        
        usuarios = User.objects.filter(email=email, is_active=True)
        if usuarios.exists():
            usuario = usuarios.first()
            token = PasswordResetToken.objects.create(usuario=usuario)
            
            reset_url = request.build_absolute_uri(f'/reset/{token.token}/')
            
            try:
                send_mail(
                    subject='Recuperar Contraseña - HotelSystem',
                    message=f'Hola {usuario.username},\n\nHaz clic en el siguiente enlace para restablecer tu contraseña:\n\n{reset_url}\n\nEste enlace expira en 1 hora.\n\nSi no solicitaste este cambio, ignora este mensaje.',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False,
                )
                messages.success(request, f'Se envió un enlace de recuperación a {email}')
            except Exception as e:
                messages.error(request, f'Error al enviar el email: {str(e)}')
        else:
            messages.info(request, 'Si el correo existe en nuestro sistema, recibirás un enlace de recuperación.')
        
        registrar_auditoria(
            usuario=None,
            accion="Solicitud Recuperación Contraseña",
            registro_id=None,
            tabla_afectada="auth_user",
            observacion=f"Email: {email}"
        )
        
        return redirect('login')
    
    return render(request, 'recuperar_contrasena.html')


@login_required
def reset_confirmar(request, token):
    """Vista para confirmar restablecimiento de contraseña"""
    from reportes.models import PasswordResetToken
    from django.contrib.auth.models import User
    
    try:
        token_obj = PasswordResetToken.objects.get(token=token)
    except PasswordResetToken.DoesNotExist:
        messages.error(request, 'Token de recuperación inválido.')
        return redirect('login')
    
    if not token_obj.esta_valido():
        messages.error(request, 'El enlace de recuperación ha expirado o ya fue usado.')
        return redirect('login')
    
    if request.method == 'POST':
        nueva_password = request.POST.get('nueva_password')
        confirmar_password = request.POST.get('confirmar_password')
        
        if nueva_password != confirmar_password:
            messages.error(request, 'Las contraseñas no coinciden.')
            return redirect('reset_confirmar', token=token)
        
        if len(nueva_password) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return redirect('reset_confirmar', token=token)
        
        if not any(c.isupper() for c in nueva_password):
            messages.error(request, 'La contraseña debe contener al menos una letra mayúscula.')
            return redirect('reset_confirmar', token=token)
        
        if not any(c.isdigit() for c in nueva_password):
            messages.error(request, 'La contraseña debe contener al menos un número.')
            return redirect('reset_confirmar', token=token)
        
        usuario = token_obj.usuario
        usuario.set_password(nueva_password)
        usuario.save()
        
        token_obj.usado = True
        token_obj.save()
        
        registrar_auditoria(
            usuario=usuario,
            accion="Contraseña Restablecida",
            registro_id=usuario.id,
            tabla_afectada="auth_user",
            estado_nuevo="Contraseña restablecida mediante token"
        )
        
        messages.success(request, 'Contraseña restablecida correctamente. Inicia sesión.')
        return redirect('login')
    
    return render(request, 'reset_confirmar.html', {'token': token})


@login_required
def desbloquear_usuario(request, user_id):
    """Admin puede desbloquear usuarios"""
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos.')
        return redirect('dashboard')
    
    from django.contrib.auth.models import User
    usuario = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        usuario.is_active = True
        usuario.save()
        
        registrar_auditoria(
            usuario=request.user,
            accion="Usuario Desbloqueado",
            registro_id=usuario.id,
            tabla_afectada="auth_user",
            estado_anterior="Bloqueado",
            estado_nuevo="Activo",
            observacion=f"Usuario '{usuario.username}' desbloqueado por {request.user.username}"
        )
        
        messages.success(request, f'Usuario {usuario.username} desbloqueado correctamente.')
    
    return redirect('usuarios_lista')


@login_required
def sesiones_activas(request):
    """Lista sesiones activas (solo admin)"""
    from django.contrib.sessions.models import Session
    from django.contrib.auth.models import User
    
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos.')
        return redirect('dashboard')
    
    sesiones = []
    for s in Session.objects.all():
        datos = s.get_decoded()
        user_id = datos.get('_auth_user_id')
        if user_id:
            try:
                usuario = User.objects.get(id=user_id)
                sesiones.append({
                    'session_key': s.session_key,
                    'usuario': usuario,
                    'ultima_actividad': s.expire_date,
                    'ip': datos.get('ip', 'unknown'),
                })
            except User.DoesNotExist:
                pass
    
    return render(request, 'sesiones/lista.html', {'sesiones': sesiones})


@login_required
def cerrar_sesion(request, session_key):
    """Admin puede cerrar sesiones remotamente"""
    from django.contrib.sessions.models import Session
    
    if not request.user.is_superuser and not request.user.groups.filter(name='admin').exists():
        messages.error(request, 'No tienes permisos.')
        return redirect('dashboard')
    
    try:
        sesion = Session.objects.get(session_key=session_key)
        datos = sesion.get_decoded()
        user_id = datos.get('_auth_user_id')
        
        if user_id:
            from django.contrib.auth.models import User
            usuario = User.objects.get(id=user_id)
            
            sesion.delete()
            
            registrar_auditoria(
                usuario=request.user,
                accion="Sesión Cerrada Remotamente",
                registro_id=usuario.id,
                tabla_afectada="auth_user",
                observacion=f"Sesión de '{usuario.username}' cerrada por {request.user.username}"
            )
            
            messages.success(request, f'Sesión de {usuario.username} cerrada correctamente.')
    except Exception as e:
        messages.error(request, f'Error al cerrar sesión: {str(e)}')
    
    return redirect('sesiones_activas')


@login_required
def actualizar_estado_habitacion(request, hab_id):
    """
    HOT-HOS-003 – Liberar / Cambiar estado de habitación manualmente.
    Permite a recepcionistas y administradores cambiar el estado físico de una habitación
    (por ejemplo, pasarla de LIMPIEZA o MANTENIMIENTO a DISPONIBLE, o viceversa).
    No permite modificar habitaciones ocupadas de forma directa.
    """
    if not (_es_recepcionista(request.user) or _es_admin(request.user)):
        return _acceso_denegado(request)

    hab = get_object_or_404(Habitacion, id=hab_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado', '').strip().upper()
        estados_validos = [Habitacion.DISPONIBLE, Habitacion.LIMPIEZA, Habitacion.MANTENIMIENTO]

        if nuevo_estado not in estados_validos:
            messages.error(request, f'Estado inválido. Los estados permitidos son: {", ".join(estados_validos)}')
            return redirect('dashboard')

        if hab.estado == Habitacion.OCUPADA:
            messages.error(request, 'No se puede cambiar el estado de una habitación ocupada de forma directa. Debe realizar un check-out o un traslado de habitación.')
            return redirect('dashboard')

        anterior = hab.estado
        hab.estado = nuevo_estado
        hab.save()

        # Registrar auditoría
        registrar_auditoria(
            usuario=request.user,
            accion="Habitación Estado Modificado",
            registro_id=hab.id,
            tabla_afectada="hotel_habitacion",
            estado_anterior=anterior,
            estado_nuevo=nuevo_estado,
            observacion=f"Cambio manual de estado de Hab. {hab.numero} de {anterior} a {nuevo_estado} (Asignar/Liberar Habitación)"
        )

        messages.success(request, f'La Habitación {hab.numero} se actualizó correctamente a {hab.get_estado_display()}.')

    return redirect('dashboard')


@login_required
def hotel_configuracion(request):
    if not (_es_admin(request.user) or _es_recepcionista(request.user) or _es_housekeeping(request.user)):
        return _acceso_denegado(request)

    from hotel.models import Hotel
    from reportes.models import Auditoria
    
    # Seeding
    hotel = Hotel.objects.first()
    if not hotel:
        hotel = Hotel.objects.create(
            nombre="Florida",
            ruc="20548732032",
            direccion="La Victoria, Lima",
            estrellas=4,
            telefono="947327362",
            correo="contacto@hotelflorida.com",
            razon_social="Hotel Florida S.A.C.",
            hora_checkin_estandar=time(15, 0),
            hora_checkout_estandar=time(12, 0),
            permitir_early_checkin=True,
            cobrar_early_checkin=False,
            early_checkin_tipo_cargo='FIJO',
            early_checkin_monto_porcentaje=Decimal('0.00'),
            permitir_late_checkout=True,
            late_checkout_tipo_cargo='FIJO',
            late_checkout_monto_porcentaje=Decimal('50.00'),
            late_checkout_hora_maxima=time(16, 0)
        )

    # Permisos específicos para visualización
    es_admin_user = _es_admin(request.user)
    es_recep_user = _es_recepcionista(request.user)
    es_house_user = _solo_housekeeping(request.user)

    # Si es housekeeping, denegar visualización de costos e historial
    mostrar_costos = not es_house_user
    mostrar_historial = not es_house_user

    historial = []
    if mostrar_historial:
        historial = Auditoria.objects.filter(tabla_afectada="hotel_hotel", accion="Cambio de Configuración").order_by('-fecha')[:50]

    if request.method == 'POST':
        if not es_admin_user:
            return _acceso_denegado(request)

        try:
            nombre = request.POST.get('nombre', '').strip()
            razon_social = request.POST.get('razon_social', '').strip() or None
            ruc = request.POST.get('ruc', '').strip()
            direccion = request.POST.get('direccion', '').strip()
            telefono = request.POST.get('telefono', '').strip()
            correo = request.POST.get('correo', '').strip() or None

            hora_in_str = request.POST.get('hora_checkin_estandar', '').strip()
            hora_out_str = request.POST.get('hora_checkout_estandar', '').strip()

            permitir_early = request.POST.get('permitir_early_checkin') in ['true', 'on', 'True']
            costo_early_str = request.POST.get('costo_early_checkin', '0.00').strip()

            permitir_late = request.POST.get('permitir_late_checkout') in ['true', 'on', 'True']
            costo_late_str = request.POST.get('costo_late_checkout', '0.00').strip()
            late_max_str = request.POST.get('late_checkout_hora_maxima', '').strip() or None

            # Validaciones básicas
            if not nombre:
                raise ValidationError("El nombre comercial del hotel es obligatorio.")
            if not ruc:
                raise ValidationError("El RUC o documento fiscal es obligatorio.")
            if not hora_in_str:
                raise ValidationError("La hora oficial de Check-In es obligatoria.")
            if not hora_out_str:
                raise ValidationError("La hora oficial de Check-Out es obligatoria.")

            # Parsing horas
            def parse_time(t_str):
                parts = t_str.split(':')
                return time(int(parts[0]), int(parts[1]))

            hora_in = parse_time(hora_in_str)
            hora_out = parse_time(hora_out_str)

            late_max = None
            if permitir_late:
                if not late_max_str:
                    raise ValidationError("La hora máxima de Late Check-Out es obligatoria si el servicio está habilitado.")
                late_max = parse_time(late_max_str)
                if late_max <= hora_out:
                    raise ValidationError("La hora máxima de Late Check-Out debe ser posterior a la hora oficial de Check-Out.")

            # Parsing costos
            costo_early = Decimal(costo_early_str)
            costo_late = Decimal(costo_late_str)

            if costo_early < 0:
                raise ValidationError("El costo de Early Check-In no puede ser negativo.")
            if costo_late < 0:
                raise ValidationError("El costo de Late Check-Out no puede ser negativo.")

            # Mapeo a campos de base de datos
            cambios = []
            
            def check_change(field, old, new, label):
                if old != new:
                    cambios.append((field, old, new, label))

            check_change('nombre', hotel.nombre, nombre, 'Nombre Comercial')
            check_change('razon_social', hotel.razon_social, razon_social, 'Razón Social')
            check_change('ruc', hotel.ruc, ruc, 'RUC')
            check_change('direccion', hotel.direccion, direccion, 'Dirección')
            check_change('telefono', hotel.telefono, telefono, 'Teléfono')
            check_change('correo', hotel.correo, correo, 'Correo Electrónico')
            check_change('hora_checkin_estandar', hotel.hora_checkin_estandar, hora_in, 'Hora Oficial Check-In')
            check_change('hora_checkout_estandar', hotel.hora_checkout_estandar, hora_out, 'Hora Oficial Check-Out')
            check_change('permitir_early_checkin', hotel.permitir_early_checkin, permitir_early, 'Permitir Early Check-In')
            
            cobrar_early = costo_early > 0
            check_change('cobrar_early_checkin', hotel.cobrar_early_checkin, cobrar_early, 'Cobrar Early Check-In')
            check_change('early_checkin_monto_porcentaje', hotel.early_checkin_monto_porcentaje, costo_early, 'Costo Early Check-In')
            
            check_change('permitir_late_checkout', hotel.permitir_late_checkout, permitir_late, 'Permitir Late Check-Out')
            check_change('late_checkout_monto_porcentaje', hotel.late_checkout_monto_porcentaje, costo_late, 'Costo Late Check-Out')
            check_change('late_checkout_hora_maxima', hotel.late_checkout_hora_maxima, late_max, 'Hora Máxima Late Check-Out')

            # Si hay cambios, actualizar y registrar auditoría
            if cambios:
                with transaction.atomic():
                    # Actualizar valores
                    hotel.nombre = nombre
                    hotel.razon_social = razon_social
                    hotel.ruc = ruc
                    hotel.direccion = direccion
                    hotel.telefono = telefono
                    hotel.correo = correo
                    hotel.hora_checkin_estandar = hora_in
                    hotel.hora_checkout_estandar = hora_out
                    hotel.permitir_early_checkin = permitir_early
                    hotel.cobrar_early_checkin = cobrar_early
                    hotel.early_checkin_tipo_cargo = 'FIJO'
                    hotel.early_checkin_monto_porcentaje = costo_early
                    hotel.permitir_late_checkout = permitir_late
                    hotel.late_checkout_tipo_cargo = 'FIJO'
                    hotel.late_checkout_monto_porcentaje = costo_late
                    hotel.late_checkout_hora_maxima = late_max
                    hotel.save()

                    # Registrar logs individuales
                    for field, old, new, label in cambios:
                        registrar_auditoria(
                            usuario=request.user,
                            accion="Cambio de Configuración",
                            registro_id=hotel.id,
                            tabla_afectada="hotel_hotel",
                            estado_anterior=str(old) if old is not None else "",
                            estado_nuevo=str(new) if new is not None else "",
                            observacion=f"Modificó el parámetro: {label}"
                        )
                messages.success(request, "Configuración guardada y auditada correctamente.")
            else:
                messages.info(request, "No se detectaron cambios en la configuración.")

            return redirect('configuracion_hotel')
        except Exception as e:
            messages.error(request, str(e))

    return render(request, 'hotel/configuracion.html', {
        'hotel': hotel,
        'mostrar_costos': mostrar_costos,
        'mostrar_historial': mostrar_historial,
        'historial': historial,
        'es_admin': es_admin_user
    })
