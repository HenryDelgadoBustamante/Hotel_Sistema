# Lógica de agregación y exportación de reportes a Excel
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from hotel.models import Habitacion
from estancias.models import Estancia
from config.permissions import EsAdmin


class OcupacionView(APIView):
    permission_classes = [EsAdmin]

    def get(self, request):
        fecha_str = request.query_params.get('fecha')
        if fecha_str:
            from datetime import datetime
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        else:
            fecha = timezone.now().date()

        total = Habitacion.objects.count()
        ocupadas = Habitacion.objects.filter(estado=Habitacion.OCUPADA).count()
        tasa = round((ocupadas / total * 100), 1) if total > 0 else 0

        estancias_activas = Estancia.objects.filter(
            estado='ACTIVA',
            fecha_checkin__date=fecha
        ).select_related('habitacion__tipo')

        revenue_por_tipo = {}
        for estancia in estancias_activas:
            tipo = estancia.habitacion.tipo.nombre
            revenue_por_tipo[tipo] = revenue_por_tipo.get(tipo, 0) + float(estancia.precio_final)

        return Response({
            'fecha': fecha,
            'total_habitaciones': total,
            'habitaciones_ocupadas': ocupadas,
            'tasa_ocupacion': tasa,
            'revenue_por_tipo': revenue_por_tipo,
        })


import datetime
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from hotel.models import TipoHabitacion
from estancias.models import CargoEstancia
from reservas.models import Reserva

@login_required
def exportar_excel(request):
    from config import roles
    if not roles.es_admin(request.user):
        from django.shortcuts import render
        return render(request, '403.html', {'mensaje_error': 'Solo los administradores pueden exportar reportes.'}, status=403)

    hoy = timezone.now().date()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    
    if fecha_inicio_str and fecha_fin_str:
        try:
            start_date = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = hoy.replace(day=1)
            end_date = hoy
    else:
        periodo = request.GET.get('periodo', 'mes')
        if periodo == 'hoy':
            start_date = hoy
            end_date = hoy
        elif periodo == 'semana':
            start_date = hoy - datetime.timedelta(days=hoy.weekday())
            end_date = start_date + datetime.timedelta(days=6)
        elif periodo == 'anio':
            start_date = hoy.replace(month=1, day=1)
            end_date = hoy.replace(month=12, day=31)
        else: # mes
            start_date = hoy.replace(day=1)
            next_month = start_date.replace(day=28) + datetime.timedelta(days=4)
            end_date = next_month - datetime.timedelta(days=next_month.day)

    dias_actual = (end_date - start_date).days + 1
    
    prev_end = start_date - datetime.timedelta(days=1)
    prev_start = prev_end - datetime.timedelta(days=dias_actual - 1)
    dias_prev = (prev_end - prev_start).days + 1

    # Obtener filtros de la URL (HOT-REP-012)
    f_piso = request.GET.get('piso', '')
    f_tipo = request.GET.get('tipo_habitacion', '')
    f_estado = request.GET.get('estado', '')
    f_habitacion = request.GET.get('habitacion', '')

    def get_period_stats(start, end, num_days):
        rooms_q = Habitacion.objects.all()
        estancias = Estancia.objects.filter(fecha_checkin__date__gte=start, fecha_checkin__date__lte=end)
        reservas = Reserva.objects.filter(fecha_entrada__gte=start, fecha_entrada__lte=end)
        cargos = CargoEstancia.objects.filter(fecha__date__gte=start, fecha__date__lte=end)
        
        if f_piso:
            rooms_q = rooms_q.filter(piso=f_piso)
            estancias = estancias.filter(habitacion__piso=f_piso)
            reservas = reservas.filter(habitacion__piso=f_piso)
            cargos = cargos.filter(estancia__habitacion__piso=f_piso)
        if f_tipo:
            rooms_q = rooms_q.filter(tipo_id=f_tipo)
            estancias = estancias.filter(habitacion__tipo_id=f_tipo)
            reservas = reservas.filter(habitacion__tipo_id=f_tipo)
            cargos = cargos.filter(estancia__habitacion__tipo_id=f_tipo)
        if f_estado:
            rooms_q = rooms_q.filter(estado=f_estado)
            if f_estado in ['ACTIVA', 'FINALIZADA', 'CANCELADA']:
                estancias = estancias.filter(estado=f_estado)
            else:
                estancias = estancias.filter(habitacion__estado=f_estado)
            if f_estado in ['PENDIENTE', 'CONFIRMADA', 'CHECKIN', 'CHECKOUT', 'CANCELADA']:
                reservas = reservas.filter(estado=f_estado)
            cargos = cargos.filter(estancia__habitacion__estado=f_estado)
        if f_habitacion:
            rooms_q = rooms_q.filter(id=f_habitacion)
            estancias = estancias.filter(habitacion_id=f_habitacion)
            reservas = reservas.filter(habitacion_id=f_habitacion)
            cargos = cargos.filter(estancia__habitacion_id=f_habitacion)
            
        total_habs = rooms_q.count()
        rooms_maint = rooms_q.filter(estado='MANTENIMIENTO').count()
        hab_disponibles = max(1, (total_habs - rooms_maint) * num_days)
        
        noches = 0
        for e in estancias:
            if e.reserva.modalidad == 'DIA':
                noches += max(1, (e.reserva.fecha_salida - e.reserva.fecha_entrada).days)
            else:
                noches += 1
                
        ingresos_totales = cargos.aggregate(total=models.Sum('monto'))['total'] or 0.0
        ingresos_hab = cargos.filter(tipo='HABITACION').aggregate(total=models.Sum('monto'))['total'] or 0.0
        
        huespedes = estancias.values('reserva__huesped').distinct().count()
        ocupacion = (noches / hab_disponibles) if hab_disponibles > 0 else 0.0
        adr = float(ingresos_hab) / noches if noches > 0 else 0.0
        revpar = float(ingresos_hab) / hab_disponibles if hab_disponibles > 0 else 0.0
        
        reservas_canceladas = reservas.filter(estado='CANCELADA').count()
        tasa_cancelacion = (reservas_canceladas / reservas.count()) if reservas.count() > 0 else 0.0
        
        return {
            'ingresos_totales': float(ingresos_totales),
            'ocupacion': float(ocupacion),
            'huespedes': huespedes,
            'noches_vendidas': noches,
            'revpar': float(revpar),
            'adr': float(adr),
            'tasa_cancelacion': float(tasa_cancelacion)
        }

    stats_actual = get_period_stats(start_date, end_date, dias_actual)
    stats_prev = get_period_stats(prev_start, prev_end, dias_prev)

    wb = Workbook()
    
    font_title = Font(name='Calibri', size=14, bold=True, color='1E293B')
    font_subtitle = Font(name='Calibri', size=11, italic=True, color='64748b')
    font_header = Font(name='Calibri', size=11, bold=True, color='1E293B')
    fill_header = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    format_currency = 'S/. #,##0.00'
    format_percentage = '0.0%'
    format_integer = '#,##0'

    # --- Hoja 1: Resumen General ---
    ws1 = wb.active
    ws1.title = "Resumen general"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1['A1'] = "REPORTE GENERAL DE RENDIMIENTO HOTELERO"
    ws1['A1'].font = font_title
    ws1['A2'] = f"Período: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}  |  Comparado con: {prev_start.strftime('%d/%m/%Y')} al {prev_end.strftime('%d/%m/%Y')}"
    ws1['A2'].font = font_subtitle
    
    headers1 = ['Métrica', 'Período Actual', 'Período Anterior', '% Variación']
    ws1.append([])
    ws1.append(headers1)
    
    for col_idx in range(1, 5):
        cell = ws1.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    def get_var_fraction(act, prev):
        if prev == 0:
            return 1.0 if act > 0 else 0.0
        return (act - prev) / prev

    metrics = [
        ('Ingresos Totales', stats_actual['ingresos_totales'], stats_prev['ingresos_totales'], get_var_fraction(stats_actual['ingresos_totales'], stats_prev['ingresos_totales']), format_currency),
        ('Nivel de Ocupación', stats_actual['ocupacion'], stats_prev['ocupacion'], get_var_fraction(stats_actual['ocupacion'], stats_prev['ocupacion']), format_percentage),
        ('Huéspedes Atendidos', stats_actual['huespedes'], stats_prev['huespedes'], get_var_fraction(stats_actual['huespedes'], stats_prev['huespedes']), format_integer),
        ('Noches Vendidas', stats_actual['noches_vendidas'], stats_prev['noches_vendidas'], get_var_fraction(stats_actual['noches_vendidas'], stats_prev['noches_vendidas']), format_integer),
        ('Rendimiento (x Habitación) - RevPAR', stats_actual['revpar'], stats_prev['revpar'], get_var_fraction(stats_actual['revpar'], stats_prev['revpar']), format_currency),
        ('Precio Promedio Vendido - ADR', stats_actual['adr'], stats_prev['adr'], get_var_fraction(stats_actual['adr'], stats_prev['adr']), format_currency),
        ('Tasa de Cancelación', stats_actual['tasa_cancelacion'], stats_prev['tasa_cancelacion'], get_var_fraction(stats_actual['tasa_cancelacion'], stats_prev['tasa_cancelacion']), format_percentage),
    ]

    for m_label, act_val, prev_val, var_val, num_fmt in metrics:
        row_data = [m_label, act_val, prev_val, var_val]
        ws1.append(row_data)
        curr_row = ws1.max_row
        
        cell_m = ws1.cell(row=curr_row, column=1)
        cell_m.alignment = align_left
        cell_m.border = border_thin
        
        cell_a = ws1.cell(row=curr_row, column=2)
        cell_a.number_format = num_fmt
        cell_a.alignment = align_right
        cell_a.border = border_thin
        
        cell_p = ws1.cell(row=curr_row, column=3)
        cell_p.number_format = num_fmt
        cell_p.alignment = align_right
        cell_p.border = border_thin
        
        cell_v = ws1.cell(row=curr_row, column=4)
        cell_v.number_format = '+0.0%;-0.0%;0.0%'
        cell_v.alignment = align_right
        cell_v.border = border_thin

    # --- Hoja 2: Estancias del período ---
    ws2 = wb.create_sheet(title="Estancias del período")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2['A1'] = "DETALLE DE ESTANCIAS"
    ws2['A1'].font = font_title
    ws2['A2'] = f"Período: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
    ws2['A2'].font = font_subtitle
    ws2.append([])
    
    headers2 = ['N°', 'Huésped', 'DNI', 'Habitación', 'Tipo', 'Check-in', 'Check-out', 'Noches', 'Tarifa/noche', 'Total', 'Estado']
    ws2.append(headers2)
    
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    estancias_qs = Estancia.objects.filter(
        fecha_checkin__date__gte=start_date, fecha_checkin__date__lte=end_date
    ).select_related('reserva__huesped', 'habitacion__tipo').order_by('-fecha_checkin')

    if f_piso:
        estancias_qs = estancias_qs.filter(habitacion__piso=f_piso)
    if f_tipo:
        estancias_qs = estancias_qs.filter(habitacion__tipo_id=f_tipo)
    if f_estado:
        if f_estado in ['ACTIVA', 'FINALIZADA', 'CANCELADA']:
            estancias_qs = estancias_qs.filter(estado=f_estado)
        else:
            estancias_qs = estancias_qs.filter(habitacion__estado=f_estado)
    if f_habitacion:
        estancias_qs = estancias_qs.filter(habitacion_id=f_habitacion)

    for idx, e in enumerate(estancias_qs, start=1):
        if e.reserva.modalidad == 'DIA':
            noches = max(1, (e.reserva.fecha_salida - e.reserva.fecha_entrada).days)
        else:
            noches = 1
            
        tarifa_noche = float(e.precio_final) / noches if noches > 0 else float(e.precio_final)
        
        checkin_val = timezone.localtime(e.fecha_checkin).replace(tzinfo=None)
        checkout_val = timezone.localtime(e.fecha_checkout).replace(tzinfo=None) if e.fecha_checkout else "-"
        
        row_data = [
            idx,
            e.reserva.huesped.nombre_completo,
            e.reserva.huesped.num_doc,
            f"Hab. {e.habitacion.numero}",
            e.habitacion.tipo.nombre,
            checkin_val,
            checkout_val,
            noches,
            tarifa_noche,
            float(e.precio_final),
            e.estado
        ]
        ws2.append(row_data)
        curr_row = ws2.max_row
        
        ws2.cell(row=curr_row, column=1).alignment = align_center
        ws2.cell(row=curr_row, column=2).alignment = align_left
        ws2.cell(row=curr_row, column=3).alignment = align_center
        ws2.cell(row=curr_row, column=4).alignment = align_center
        ws2.cell(row=curr_row, column=5).alignment = align_left
        
        c_in = ws2.cell(row=curr_row, column=6)
        if isinstance(checkin_val, datetime.datetime):
            c_in.number_format = 'dd/mm/yyyy hh:mm'
        c_in.alignment = align_center
        
        c_out = ws2.cell(row=curr_row, column=7)
        if isinstance(checkout_val, datetime.datetime):
            c_out.number_format = 'dd/mm/yyyy hh:mm'
        c_out.alignment = align_center
        
        c_n = ws2.cell(row=curr_row, column=8)
        c_n.number_format = format_integer
        c_n.alignment = align_center
        
        c_t = ws2.cell(row=curr_row, column=9)
        c_t.number_format = format_currency
        c_t.alignment = align_right
        
        c_tot = ws2.cell(row=curr_row, column=10)
        c_tot.number_format = format_currency
        c_tot.alignment = align_right
        
        ws2.cell(row=curr_row, column=11).alignment = align_center
        
        for col_idx in range(1, len(headers2) + 1):
            ws2.cell(row=curr_row, column=col_idx).border = border_thin

    # --- Hoja 3: Ocupación por tipo de habitación ---
    ws3 = wb.create_sheet(title="Ocupación por tipo")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3['A1'] = "OCUPACIÓN POR TIPO DE HABITACIÓN"
    ws3['A1'].font = font_title
    ws3['A2'] = f"Período: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
    ws3['A2'].font = font_subtitle
    ws3.append([])
    
    headers3 = ['Tipo de Habitación', 'Total Hab. Disponibles', 'Hab. Ocupadas (Noches)', '% Ocupación', 'Ingresos Generados']
    ws3.append(headers3)
    
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    tipos_hab = TipoHabitacion.objects.all()
    if f_tipo:
        tipos_hab = tipos_hab.filter(id=f_tipo)
        
    for tipo in tipos_hab:
        rooms_tipo_q = Habitacion.objects.filter(tipo=tipo)
        if f_piso:
            rooms_tipo_q = rooms_tipo_q.filter(piso=f_piso)
        if f_habitacion:
            rooms_tipo_q = rooms_tipo_q.filter(id=f_habitacion)
            
        total_rooms = rooms_tipo_q.count()
        hab_disponibles_tipo = total_rooms * dias_actual
        
        noches_tipo = 0
        estancias_tipo = estancias_qs.filter(habitacion__tipo=tipo)
        for e in estancias_tipo:
            if e.reserva.modalidad == 'DIA':
                noches_tipo += max(1, (e.reserva.fecha_salida - e.reserva.fecha_entrada).days)
            else:
                noches_tipo += 1
                
        pct_ocupacion = (noches_tipo / hab_disponibles_tipo) if hab_disponibles_tipo > 0 else 0.0
        
        cargos_tipo_q = CargoEstancia.objects.filter(
            tipo='HABITACION',
            estancia__habitacion__tipo=tipo,
            fecha__date__gte=start_date,
            fecha__date__lte=end_date
        )
        if f_piso:
            cargos_tipo_q = cargos_tipo_q.filter(estancia__habitacion__piso=f_piso)
        if f_habitacion:
            cargos_tipo_q = cargos_tipo_q.filter(estancia__habitacion_id=f_habitacion)
            
        ingresos_tipo = cargos_tipo_q.aggregate(total=models.Sum('monto'))['total'] or 0.0
        
        row_data = [
            tipo.nombre,
            hab_disponibles_tipo,
            noches_tipo,
            pct_ocupacion,
            float(ingresos_tipo)
        ]
        ws3.append(row_data)
        curr_row = ws3.max_row
        
        ws3.cell(row=curr_row, column=1).alignment = align_left
        
        c_disp = ws3.cell(row=curr_row, column=2)
        c_disp.number_format = format_integer
        c_disp.alignment = align_center
        
        c_ocup = ws3.cell(row=curr_row, column=3)
        c_ocup.number_format = format_integer
        c_ocup.alignment = align_center
        
        c_pct = ws3.cell(row=curr_row, column=4)
        c_pct.number_format = format_percentage
        c_pct.alignment = align_right
        
        c_ing = ws3.cell(row=curr_row, column=5)
        c_ing.number_format = format_currency
        c_ing.alignment = align_right
        
        for col_idx in range(1, len(headers3) + 1):
            ws3.cell(row=curr_row, column=col_idx).border = border_thin

    # --- Hoja 4: Ingresos por mes (últimos 12 meses) ---
    ws4 = wb.create_sheet(title="Ingresos mensuales")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4['A1'] = "HISTÓRICO DE INGRESOS MENSUALES (ÚLTIMOS 12 MESES)"
    ws4['A1'].font = font_title
    ws4.append([])
    
    headers4 = ['Mes', 'Año', 'N° Estancias', 'N° Huéspedes', 'Ingresos Totales', 'RevPAR del Mes']
    ws4.append(headers4)
    
    for col_idx in range(1, len(headers4) + 1):
        cell = ws4.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    months_list = []
    current = hoy.replace(day=1)
    for _ in range(12):
        months_list.append((current.year, current.month))
        if current.month == 1:
            current = current.replace(year=current.year - 1, month=12)
        else:
            current = current.replace(month=current.month - 1)
    months_list.reverse()

    nombre_meses = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
        7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }

    rooms_all = Habitacion.objects.all()
    if f_piso: rooms_all = rooms_all.filter(piso=f_piso)
    if f_tipo: rooms_all = rooms_all.filter(tipo_id=f_tipo)
    if f_estado: rooms_all = rooms_all.filter(estado=f_estado)
    if f_habitacion: rooms_all = rooms_all.filter(id=f_habitacion)
    total_habitaciones = rooms_all.count()

    for year, month in months_list:
        start_m = datetime.date(year, month, 1)
        if month == 12:
            end_m = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            end_m = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
            
        dias_m = (end_m - start_m).days + 1
        hab_disponibles_m = total_habitaciones * dias_m
        
        estancias_m = Estancia.objects.filter(fecha_checkin__date__gte=start_m, fecha_checkin__date__lte=end_m)
        if f_piso: estancias_m = estancias_m.filter(habitacion__piso=f_piso)
        if f_tipo: estancias_m = estancias_m.filter(habitacion__tipo_id=f_tipo)
        if f_estado:
            if f_estado in ['ACTIVA', 'FINALIZADA', 'CANCELADA']:
                estancias_m = estancias_m.filter(estado=f_estado)
            else:
                estancias_m = estancias_m.filter(habitacion__estado=f_estado)
        if f_habitacion: estancias_m = estancias_m.filter(habitacion_id=f_habitacion)
        
        num_estancias = estancias_m.count()
        num_huespedes = estancias_m.values('reserva__huesped').distinct().count()
        
        cargos_m = CargoEstancia.objects.filter(fecha__date__gte=start_m, fecha__date__lte=end_m)
        if f_piso: cargos_m = cargos_m.filter(estancia__habitacion__piso=f_piso)
        if f_tipo: cargos_m = cargos_m.filter(estancia__habitacion__tipo_id=f_tipo)
        if f_estado: cargos_m = cargos_m.filter(estancia__habitacion__estado=f_estado)
        if f_habitacion: cargos_m = cargos_m.filter(estancia__habitacion_id=f_habitacion)
        
        ingresos_m = cargos_m.aggregate(total=models.Sum('monto'))['total'] or 0.0
        ingresos_hab_m = cargos_m.filter(tipo='HABITACION').aggregate(total=models.Sum('monto'))['total'] or 0.0
        
        revpar_m = float(ingresos_hab_m) / hab_disponibles_m if hab_disponibles_m > 0 else 0.0
        
        row_data = [
            nombre_meses[month],
            year,
            num_estancias,
            num_huespedes,
            float(ingresos_m),
            revpar_m
        ]
        ws4.append(row_data)
        curr_row = ws4.max_row
        
        ws4.cell(row=curr_row, column=1).alignment = align_left
        ws4.cell(row=curr_row, column=2).alignment = align_center
        
        c_est = ws4.cell(row=curr_row, column=3)
        c_est.number_format = format_integer
        c_est.alignment = align_center
        
        c_hues = ws4.cell(row=curr_row, column=4)
        c_hues.number_format = format_integer
        c_hues.alignment = align_center
        
        c_ing = ws4.cell(row=curr_row, column=5)
        c_ing.number_format = format_currency
        c_ing.alignment = align_right
        
        c_rev = ws4.cell(row=curr_row, column=6)
        c_rev.number_format = format_currency
        c_rev.alignment = align_right
        
        for col_idx in range(1, len(headers4) + 1):
            ws4.cell(row=curr_row, column=col_idx).border = border_thin

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = cell.value
                if val is not None:
                    if isinstance(val, datetime.date):
                        s = val.strftime('%d/%m/%Y')
                    elif isinstance(val, datetime.datetime):
                        s = val.strftime('%d/%m/%Y %H:%M')
                    else:
                        s = str(val)
                    max_len = max(max_len, len(s))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"reporte_hotel_{hoy.strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

